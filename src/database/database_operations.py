import os
import shutil
import glob

import uuid as _uuid
import json
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional

from sqlalchemy.exc import IntegrityError, SQLAlchemyError, NoResultFound
from sqlalchemy.inspection import inspect
from sqlalchemy import cast, Date, func, case
from passlib.context import CryptContext

from .database_create import SessionLocal
from .database_tables import (
    Departments_Roles, Admins, Employees, LoginHistory, Managers,
    Projects, SRS_Documents, ProjectTimeline, ProjectAssignments, MilestoneAssignments,
    DeveloperTasks, ContentCreatorTasks, ProjectExpenses, Attendance, LeaveRequest, ProjectPayments,
    ChecklistTemplate, ProjectChecklistState, AuditLog, InAppNotification, ClientReceivables
)
# ==========================================
#              LOGGING SETUP
# ==========================================
# Professional logging setup for granular debugging.
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(name)s | %(message)s'
)
logger = logging.getLogger("YanaDB_Operations")

# Setup password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

class DatabaseOperations:

    def __init__(self):
        # We use the factory directly so we can generate isolated sessions per request
        self.SessionLocal = SessionLocal

    # ==========================================
    #             UTILITY METHODS
    # ==========================================

    def model_to_dict(self, obj) -> dict:
        """Safely converts a SQLAlchemy model instance to a Python dictionary."""
        if not obj:
            return {}
        res = {c.key: getattr(obj, c.key) for c in inspect(obj).mapper.column_attrs}
        # Decrypt Aadhaar and PAN if present in the model dictionary
        if "adhar_number" in res and res["adhar_number"]:
            from .encryption import decrypt_data
            res["adhar_number"] = decrypt_data(res["adhar_number"])
            res["aadhaar_number"] = res["adhar_number"]
        elif "adhar_number" in res:
            res["aadhaar_number"] = res["adhar_number"]
        if "pan_number" in res and res["pan_number"]:
            from .encryption import decrypt_data
            res["pan_number"] = decrypt_data(res["pan_number"])
        return res

    def write_audit_log(self, user_id: str, action: str, target_id: str = None, details: dict = None) -> str:
        """Writes an audit log entry to the database."""
        with self.SessionLocal() as session:
            try:
                details_str = json.dumps(details) if details else None
                new_log = AuditLog(
                    user_id=user_id,
                    action=action,
                    target_id=target_id,
                    details=details_str
                )
                session.add(new_log)
                session.commit()
                session.refresh(new_log)
                return json.dumps(self.model_to_dict(new_log), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("write_audit_log", e)

    def _parse_datetime(self, val: str) -> Optional[datetime]:
        """
        Attempts to parse a string representation of datetime into a Python datetime object.
        Supports various standard formats.
        """
        if not val or val.strip() in ("N/A", "None", "nan", "NaN", "NaT"):
            return None
        val_str = val.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(val_str, fmt)
            except ValueError:
                continue
        # ISO fallback
        try:
            return datetime.fromisoformat(val_str.replace(" ", "T"))
        except ValueError:
            return None

    def _sanitize_payload(self, data: dict) -> dict:
        """
        Cleanses incoming project/attendance/payment payloads.
        Converts Pandas/NumPy NaN, NaT, and null-like objects to Python None,
        and automatically converts string-casted Datetime fields to datetime objects,
        preventing SQLite parameter binding errors.
        """
        import math
        from datetime import datetime as dt_class
        sanitized = {}
        datetime_keys = {
            "created_at", "updated_at", "payment_date", "login_timestamp", "date",
            "check_in_time", "check_out_time", "expected_start", "expected_end",
            "actual_start", "actual_end"
        }
        for k, v in data.items():
            if v is None:
                sanitized[k] = None
                continue
                
            v_str = str(v).strip()
            
            # Check for Pandas NaT or nan or empty values
            if v_str in ("NaT", "nan", "NaN", "") or type(v).__name__ == "NaTType":
                sanitized[k] = None
                continue
                
            if isinstance(v, float) and math.isnan(v):
                sanitized[k] = None
                continue
                
            # Handle Pandas Timestamp / Python datetime objects
            if type(v).__name__ == "Timestamp" or hasattr(v, "to_pydatetime") or isinstance(v, dt_class):
                py_dt = v.to_pydatetime() if hasattr(v, "to_pydatetime") else v
                if k in {"start_date", "end_date"}:
                    sanitized[k] = py_dt.strftime("%Y-%m-%d")
                else:
                    sanitized[k] = py_dt
                continue
                
            # Handle strings for datetime or start/end date
            if isinstance(v, str):
                if k in datetime_keys or k in {"created_at", "updated_at"}:
                    parsed = self._parse_datetime(v)
                    sanitized[k] = parsed
                elif k in {"start_date", "end_date"}:
                    # Clean up date strings like "2026-02-21 00:00:00" to "2026-02-21"
                    if len(v) >= 10 and v[4] == "-" and v[7] == "-":
                        sanitized[k] = v[:10]
                    else:
                        sanitized[k] = v
                else:
                    sanitized[k] = v
            else:
                sanitized[k] = v
        return sanitized

    def _handle_error(self, operation: str, exception: Exception, context: dict = None) -> str:
        """
        Centralized, professional error handler. 
        Logs the full stack trace securely on the server, while returning a safe JSON structure to the client.
        """
        context_str = json.dumps(context, default=str) if context else "No context provided"
        logger.error(f"DB_ERROR in '{operation}' | Context: {context_str} | Err: {str(exception)}", exc_info=True)
        
        if isinstance(exception, IntegrityError):
            return json.dumps({
                "error": "Database integrity violation. A required link (foreign key) is missing, or a unique constraint (like an existing username) was violated.",
                "details": str(exception.orig)
            })
        elif isinstance(exception, SQLAlchemyError):
            return json.dumps({
                "critical_error": "A deep database transaction error occurred.",
                "details": str(exception)
            })
        else:
            return json.dumps({
                "critical_error": "An unexpected application error occurred.",
                "details": str(exception)
            })

    # ==========================================
    #              MANAGERS
    # ==========================================

    def add_manager(self, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                new_manager = Managers(**data)
                session.add(new_manager)
                session.commit()
                session.refresh(new_manager)
                return json.dumps(self.model_to_dict(new_manager), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("add_manager", e, context={"name": data.get("name")})

    def get_all_managers(self) -> str:
        with self.SessionLocal() as session:
            try:
                managers = session.query(Managers).all()
                return json.dumps([self.model_to_dict(m) for m in managers], indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_all_managers", e)

    # ==========================================
    #        DEPARTMENTS & ROLES (PHASE 2)
    # ==========================================

    def add_department_role(self, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                dept_role = Departments_Roles(**data)
                session.add(dept_role)
                session.commit()
                session.refresh(dept_role)
                logger.info(f"Created new Department/Role: {dept_role.department_name} - {dept_role.role_name}")
                return json.dumps(self.model_to_dict(dept_role), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("add_department_role", e, context=data)

    def get_all_departments_roles(self) -> str:
        with self.SessionLocal() as session:
            try:
                roles = session.query(Departments_Roles).all()
                return json.dumps([self.model_to_dict(r) for r in roles], indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_all_departments_roles", e)

    # ==========================================
    #        ADMINS & EMPLOYEES (PHASE 2)
    # ==========================================

    def add_admin(self, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                default_password = data.get("password") or "AdminPass123!"
                data['password'] = get_password_hash(default_password)
                
                new_admin = Admins(**data)
                session.add(new_admin)
                session.commit()
                session.refresh(new_admin)
                
                # Automatically create Employee Profile and insert into Managers table if access_level is ManagerAdmin
                if new_admin.access_level == "ManagerAdmin":
                    # Create Employees profile
                    existing_emp = session.query(Employees).filter_by(username=new_admin.username).first()
                    if not existing_emp:
                        # Find a role with "manager" in name, or default role
                        manager_role = session.query(Departments_Roles).filter(Departments_Roles.role_name.ilike("%manager%")).first()
                        role_id = manager_role.id if manager_role else None
                        
                        # Generate unique email if N/A
                        email_val = new_admin.email if new_admin.email and new_admin.email != "N/A" else f"{new_admin.username}@yanatech.com"
                        
                        new_emp = Employees(
                            username=new_admin.username,
                            password=new_admin.password,
                            # plain_password=default_password,
                            full_name=new_admin.full_name if new_admin.full_name and new_admin.full_name != "N/A" else new_admin.username,
                            email=email_val,
                            role_id=role_id,
                            is_active=True
                        )
                        session.add(new_emp)
                    
                    # Insert into Managers registry
                    mgr_name = new_admin.full_name if new_admin.full_name and new_admin.full_name != "N/A" else new_admin.username
                    existing_mgr = session.query(Managers).filter_by(name=mgr_name).first()
                    if not existing_mgr:
                        new_mgr = Managers(name=mgr_name)
                        session.add(new_mgr)
                    
                    session.commit()
                
                # Strip password from return dict for security
                res_dict = self.model_to_dict(new_admin)
                res_dict.pop('password', None)
                res_dict['generated_password'] = default_password # Expose plaintext once
                
                logger.info(f"Created new Admin: {new_admin.username}")
                return json.dumps(res_dict, indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("add_admin", e, context={"username": data.get("username")})

    def add_employee(self, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                # 1. Username Auto-Generation Logic (if missing)
                full_name = data.get("full_name", "employee").strip()
                if not full_name or full_name.lower() == "undefined":
                    full_name = "employee"

                if not data.get("username"):
                    base_username = full_name.lower().replace(" ", ".")
                    username = base_username
                    counter = 1
                    # Check for uniqueness across Employees AND Admins
                    while session.query(Employees).filter_by(username=username).first() or \
                          session.query(Admins).filter_by(username=username).first():
                        username = f"{base_username}{counter}"
                        counter += 1
                    data["username"] = username

                # 2. Password Handling
                default_password = data.get("password") or "YanaUser123!"
                data["password"] = get_password_hash(default_password)
                # data["plain_password"] = default_password

                # 2.5 Salary -> Hourly Cost Auto-calculation
                salary_val = float(data.get("salary") or 0.0)
                data["hourly_cost_rate"] = salary_val / (26 * 8)

                # Remove extra fields not in table
                data.pop("department", None)

                # Map aadhaar_number to adhar_number if present
                if "aadhaar_number" in data:
                    data["adhar_number"] = data.pop("aadhaar_number")

                # Encrypt Aadhaar & PAN before database write
                from .encryption import encrypt_data
                if "adhar_number" in data and data["adhar_number"]:
                    data["adhar_number"] = encrypt_data(data["adhar_number"])
                if "pan_number" in data and data["pan_number"]:
                    data["pan_number"] = encrypt_data(data["pan_number"])

                # 3. Insert Database Record
                new_employee = Employees(**data)
                session.add(new_employee)
                session.commit()
                session.refresh(new_employee)
                
                res_dict = self.model_to_dict(new_employee)
                res_dict['generated_password'] = default_password # Return plain text ONCE for the admin to copy
                
                logger.info(f"Created new Employee: {new_employee.username}")
                return json.dumps(res_dict, indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("add_employee", e, context={"full_name": data.get("full_name")})

    def get_all_employees(self) -> str:
        with self.SessionLocal() as session:
            try:
                employees = session.query(Employees).all()
                results = []
                for emp in employees:
                    emp_dict = self.model_to_dict(emp)
                    emp_dict.pop('password', None)
                    results.append(emp_dict)
                return json.dumps(results, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_all_employees", e)

    def edit_employee(self, employee_id: str, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                employee = session.query(Employees).filter_by(id=employee_id).first()
                if not employee:
                    return json.dumps({"error": f"Employee {employee_id} not found."})

                # Map aadhaar_number to adhar_number if present
                if "aadhaar_number" in data:
                    data["adhar_number"] = data.pop("aadhaar_number")

                from .encryption import encrypt_data
                for key, value in data.items():
                    if hasattr(employee, key):
                        if key == "password":
                            value = get_password_hash(value)
                        elif key in ("adhar_number", "pan_number") and value:
                            value = encrypt_data(value)
                        setattr(employee, key, value)
                
                # Auto-recalculate hourly cost rate when salary is edited
                if "salary" in data:
                    employee.hourly_cost_rate = float(data["salary"] or 0.0) / (26 * 8)
                
                session.commit()
                session.refresh(employee)
                res = self.model_to_dict(employee)
                res.pop('password', None)
                return json.dumps(res, indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("edit_employee", e, context={"employee_id": employee_id})

    # ==========================================
    #     TASK INTERCEPTORS & CALCULATIONS
    # ==========================================

    def add_developer_task(self, data: dict) -> str:
        """
        Calculation Engine Trigger: Intercepts developer timesheet, auto-calculates 
        costs and profits based on Employee's preset hourly rates.
        """
        with self.SessionLocal() as session:
            try:
                employee_id = data.get("employee_id")
                if not employee_id:
                    return json.dumps({"error": "employee_id is required for Developer Tasks."})
                
                # --- ATTENDANCE CHECK-IN/CHECK-OUT ENFORCEMENT ---
                task_date_val = data.get("date")
                if not task_date_val:
                    task_date = datetime.now()
                elif isinstance(task_date_val, datetime):
                    task_date = task_date_val
                elif hasattr(task_date_val, "date"):
                    task_date = datetime.combine(task_date_val, datetime.min.time())
                elif isinstance(task_date_val, str):
                    parsed = None
                    for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                        try:
                            parsed = datetime.strptime(task_date_val.split("+")[0].split("Z")[0], fmt)
                            break
                        except ValueError:
                            continue
                    if parsed is None:
                        try:
                            parsed = datetime.fromisoformat(task_date_val)
                        except ValueError:
                            pass
                    task_date = parsed if parsed is not None else datetime.now()
                else:
                    task_date = datetime.now()

                # Normalize to date to perform date.between check
                task_cal_date = task_date.date()
                start_of_day = datetime.combine(task_cal_date, datetime.min.time())
                end_of_day = datetime.combine(task_cal_date, datetime.max.time())

                attendance = session.query(Attendance).filter(
                    Attendance.employee_id == employee_id,
                    Attendance.date.between(start_of_day, end_of_day),
                    Attendance.check_in_time.isnot(None)
                ).order_by(Attendance.check_in_time.desc()).first()

                if not attendance:
                    return json.dumps({"error": "Access Denied: You must check in before logging any tasks."})
                
                # Check if employee has any completed shift today
                has_completed_shift_today = session.query(Attendance).filter(
                    Attendance.employee_id == employee_id,
                    Attendance.date.between(start_of_day, end_of_day),
                    Attendance.check_out_time.isnot(None)
                ).first() is not None

                if attendance.check_out_time is not None:
                    # Automatically check them in for an extended/overtime shift
                    extended_attendance = Attendance(
                        employee_id=employee_id,
                        check_in_time=datetime.now(),
                        date=datetime.now(),
                        status="Checked In",
                        attendance_status="Extended Shift",
                        ip_address=attendance.ip_address,
                        notes="Auto Checked In for Extended Shift"
                    )
                    session.add(extended_attendance)
                    session.flush()
                    attendance = extended_attendance
                    has_completed_shift_today = True

                employee = session.query(Employees).filter_by(id=employee_id).first()
                if not employee:
                    return json.dumps({"error": f"Employee {employee_id} not found. Cannot calculate costs."})

                # Check 24 hours daily limit for timesheet entries
                dev_hours_today = session.query(func.sum(DeveloperTasks.hours_logged)).filter(
                    DeveloperTasks.employee_id == employee_id,
                    DeveloperTasks.date.between(start_of_day, end_of_day)
                ).scalar() or 0.0
                
                content_hours_today = session.query(func.sum(ContentCreatorTasks.hours_logged)).filter(
                    ContentCreatorTasks.employee_id == employee_id,
                    ContentCreatorTasks.date.between(start_of_day, end_of_day)
                ).scalar() or 0.0
                
                existing_hours_today = float(dev_hours_today) + float(content_hours_today)
                hours_logged = float(data.get("hours_logged", 0.0))
                
                if existing_hours_today + hours_logged > 24.0:
                    return json.dumps({"error": f"Daily limit exceeded: You have already logged {existing_hours_today} hours on this date. Logging {hours_logged} hours would exceed the 24 hours daily limit."})

                normal_limit = float(employee.working_hours or 8.0)
                if has_completed_shift_today or (existing_hours_today + hours_logged > normal_limit):
                    data["is_overtime"] = True

                # Store the parsed datetime in data
                data["date"] = task_date
                    
                project_id = data.get("project_id")
                cost_rate = float(employee.hourly_cost_rate or 0.0)
                billing_rate = float(employee.hourly_billing_rate or 0.0)

                # Intercept logic: If working on a specific project, check for custom whitelist rates
                if project_id:
                    # If it's a handover task, check assignment for the original colleague
                    check_emp_id = employee_id
                    if data.get("is_handover") and data.get("handover_for_employee_id"):
                        check_emp_id = data.get("handover_for_employee_id")

                    assignment = session.query(ProjectAssignments).filter_by(
                        project_id=project_id, 
                        employee_id=check_emp_id
                    ).first()
                    
                    if not assignment:
                        if data.get("is_handover") and data.get("handover_for_employee_id"):
                            return json.dumps({"error": "The colleague you are covering for is not assigned to this project. Access denied."})
                        return json.dumps({"error": "You are not assigned to this project. Access denied."})

                    if assignment.custom_hourly_billing is not None:
                        billing_rate = float(assignment.custom_hourly_billing)

                hours_logged = float(data.get("hours_logged", 0.0))
                data["employee_cost"] = hours_logged * cost_rate
                data["billing_amount"] = hours_logged * billing_rate
                data["profit_loss"] = data["billing_amount"] - data["employee_cost"]

                # Save Task
                new_task = DeveloperTasks(**data)
                session.add(new_task)
                
                # Check for deployment and notify/create expense
                was_deployed = data.get("was_deployed", "No")
                project_id = data.get("project_id")
                if was_deployed == "Yes" and project_id:
                    dev_deploy_count = session.query(DeveloperTasks).filter(
                        DeveloperTasks.project_id == project_id,
                        DeveloperTasks.was_deployed == "Yes"
                    ).count()
                    content_deploy_count = session.query(ContentCreatorTasks).filter(
                        ContentCreatorTasks.project_id == project_id,
                        ContentCreatorTasks.was_deployed == "Yes"
                    ).count()
                    is_first_deploy = (dev_deploy_count == 0 and content_deploy_count == 0)
                    
                    emp_name = employee.full_name if employee else "Employee"
                    t_id = data.get("ticket_id") or "N/A"
                    expense_data = ProjectExpenses(
                        project_id=project_id,
                        expense_name=f"Deployment - {t_id}",
                        amount=0.0,
                        expense_date=datetime.now().strftime("%Y-%m-%d"),
                        description=f"Auto-generated deployment expense by {emp_name} for task {t_id}."
                    )
                    session.add(expense_data)
                    
                    if is_first_deploy:
                        proj = session.query(Projects).filter_by(id=project_id).first()
                        proj_name = proj.name if proj else "Unknown Project"
                        mgr_name = proj.manager if proj else None
                        
                        admins = session.query(Admins).all()
                        for admin in admins:
                            new_notif = InAppNotification(
                                user_id=admin.id,
                                title="First Project Deployment",
                                message=f"Project '{proj_name}' has received its first deployment by {emp_name} (Task: {t_id}). Please verify the deployment.",
                                is_read=False
                            )
                            session.add(new_notif)
                
                # Check and update milestone status and actual_start date
                milestone_id = data.get("milestone_id")
                if milestone_id:
                    milestone = session.query(ProjectTimeline).filter_by(id=milestone_id).first()
                    if milestone:
                        if milestone.status == 'Pending':
                            milestone.status = 'Active'
                        if not milestone.actual_start:
                            milestone.actual_start = datetime.now()
                        
                # Update status of original handover task if handover_source_task_id is provided
                source_task_id = data.get("handover_source_task_id")
                if source_task_id:
                    orig_task = session.query(DeveloperTasks).filter_by(id=source_task_id).first()
                    if not orig_task:
                        orig_task = session.query(ContentCreatorTasks).filter_by(id=source_task_id).first()
                    if orig_task:
                        orig_task.task_status = data.get("task_status", "Completed")
                    
                    # Also update LeaveRequest if matches source_task_id
                    leave_req = session.query(LeaveRequest).filter_by(id=source_task_id).first()
                    if leave_req:
                        leave_req.handover_status = "Completed" if data.get("task_status") == "Completed" else "Pending"

                session.commit()
                session.refresh(new_task)

                
                logger.info(f"Developer Task Added | Emp: {employee.username} | Profit: ${data['profit_loss']}")
                return json.dumps(self.model_to_dict(new_task), indent=4, default=str)
            
            except ValueError as ve:
                session.rollback()
                return json.dumps({"error": "Invalid numerical data submitted for hours or rates.", "details": str(ve)})
            except Exception as e:
                session.rollback()
                return self._handle_error("add_developer_task", e, context=data)

    def add_content_creator_task(self, data: dict) -> str:
        """
        Calculation Engine Trigger: Intercepts social media task, auto-calculates 
        total content generated based on pieces.
        """
        with self.SessionLocal() as session:
            try:
                employee_id = data.get("employee_id")
                if not employee_id:
                    return json.dumps({"error": "employee_id is required for Content Tasks."})

                # --- ATTENDANCE CHECK-IN/CHECK-OUT ENFORCEMENT ---
                task_date_val = data.get("date")
                if not task_date_val:
                    task_date = datetime.now()
                elif isinstance(task_date_val, datetime):
                    task_date = task_date_val
                elif hasattr(task_date_val, "date"):
                    task_date = datetime.combine(task_date_val, datetime.min.time())
                elif isinstance(task_date_val, str):
                    parsed = None
                    for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                        try:
                            parsed = datetime.strptime(task_date_val.split("+")[0].split("Z")[0], fmt)
                            break
                        except ValueError:
                            continue
                    if parsed is None:
                        try:
                            parsed = datetime.fromisoformat(task_date_val)
                        except ValueError:
                            pass
                    task_date = parsed if parsed is not None else datetime.now()
                else:
                    task_date = datetime.now()

                # Normalize to date to perform date.between check
                task_cal_date = task_date.date()
                start_of_day = datetime.combine(task_cal_date, datetime.min.time())
                end_of_day = datetime.combine(task_cal_date, datetime.max.time())

                attendance = session.query(Attendance).filter(
                    Attendance.employee_id == employee_id,
                    Attendance.date.between(start_of_day, end_of_day),
                    Attendance.check_in_time.isnot(None)
                ).order_by(Attendance.check_in_time.desc()).first()

                if not attendance:
                    return json.dumps({"error": "Access Denied: You must check in before logging any tasks."})
                
                # Check if employee has any completed shift today
                has_completed_shift_today = session.query(Attendance).filter(
                    Attendance.employee_id == employee_id,
                    Attendance.date.between(start_of_day, end_of_day),
                    Attendance.check_out_time.isnot(None)
                ).first() is not None

                if attendance.check_out_time is not None:
                    # Automatically check them in for an extended/overtime shift
                    extended_attendance = Attendance(
                        employee_id=employee_id,
                        check_in_time=datetime.now(),
                        date=datetime.now(),
                        status="Checked In",
                        attendance_status="Extended Shift",
                        ip_address=attendance.ip_address,
                        notes="Auto Checked In for Extended Shift"
                    )
                    session.add(extended_attendance)
                    session.flush()
                    attendance = extended_attendance
                    has_completed_shift_today = True

                employee = session.query(Employees).filter_by(id=employee_id).first()
                if not employee:
                    return json.dumps({"error": f"Employee {employee_id} not found."})

                # Check 24 hours daily limit for timesheet entries
                dev_hours_today = session.query(func.sum(DeveloperTasks.hours_logged)).filter(
                    DeveloperTasks.employee_id == employee_id,
                    DeveloperTasks.date.between(start_of_day, end_of_day)
                ).scalar() or 0.0
                
                content_hours_today = session.query(func.sum(ContentCreatorTasks.hours_logged)).filter(
                    ContentCreatorTasks.employee_id == employee_id,
                    ContentCreatorTasks.date.between(start_of_day, end_of_day)
                ).scalar() or 0.0
                
                existing_hours_today = float(dev_hours_today) + float(content_hours_today)
                hours_logged = float(data.get("hours_logged", 0.0))
                
                if existing_hours_today + hours_logged > 24.0:
                    return json.dumps({"error": f"Daily limit exceeded: You have already logged {existing_hours_today} hours on this date. Logging {hours_logged} hours would exceed the 24 hours daily limit."})

                normal_limit = float(employee.working_hours or 8.0)
                if has_completed_shift_today or (existing_hours_today + hours_logged > normal_limit):
                    data["is_overtime"] = True

                # Store the parsed datetime in data
                data["date"] = task_date

                project_id = data.get("project_id")
                cost_rate = float(employee.hourly_cost_rate or 0.0)
                billing_rate = float(employee.hourly_billing_rate or 0.0)

                # Intercept logic: If working on a specific project, check for custom whitelist rates
                if project_id:
                    # If it's a handover task, check assignment for the original colleague
                    check_emp_id = employee_id
                    if data.get("is_handover") and data.get("handover_for_employee_id"):
                        check_emp_id = data.get("handover_for_employee_id")

                    assignment = session.query(ProjectAssignments).filter_by(
                        project_id=project_id, 
                        employee_id=check_emp_id
                    ).first()
                    
                    if not assignment:
                        if data.get("is_handover") and data.get("handover_for_employee_id"):
                            return json.dumps({"error": "The colleague you are covering for is not assigned to this project. Access denied."})
                        return json.dumps({"error": "You are not assigned to this project. Access denied."})

                    if assignment.custom_hourly_billing is not None:
                        billing_rate = float(assignment.custom_hourly_billing)

                hours_logged = float(data.get("hours_logged", 0.0))
                data["employee_cost"] = hours_logged * cost_rate
                data["billing_amount"] = hours_logged * billing_rate
                data["profit_loss"] = data["billing_amount"] - data["employee_cost"]

                # CONTENT AGGREGATION AUTO-CALCULATION
                reels = int(data.get("reels_count", 0))
                long_videos = int(data.get("long_video_count", 0))
                posters = int(data.get("poster_count", 0))
                
                data["total_content"] = reels + long_videos + posters

                # Save Task
                new_task = ContentCreatorTasks(**data)
                session.add(new_task)
                
                # Check for deployment and notify/create expense
                was_deployed = data.get("was_deployed", "No")
                project_id = data.get("project_id")
                if was_deployed == "Yes" and project_id:
                    dev_deploy_count = session.query(DeveloperTasks).filter(
                        DeveloperTasks.project_id == project_id,
                        DeveloperTasks.was_deployed == "Yes"
                    ).count()
                    content_deploy_count = session.query(ContentCreatorTasks).filter(
                        ContentCreatorTasks.project_id == project_id,
                        ContentCreatorTasks.was_deployed == "Yes"
                    ).count()
                    is_first_deploy = (dev_deploy_count == 0 and content_deploy_count == 0)
                    
                    emp_name = employee.full_name if employee else "Employee"
                    t_id = data.get("ticket_id") or "N/A"
                    expense_data = ProjectExpenses(
                        project_id=project_id,
                        expense_name=f"Deployment - {t_id}",
                        amount=0.0,
                        expense_date=datetime.now().strftime("%Y-%m-%d"),
                        description=f"Auto-generated deployment expense by {emp_name} for task {t_id}."
                    )
                    session.add(expense_data)
                    
                    if is_first_deploy:
                        proj = session.query(Projects).filter_by(id=project_id).first()
                        proj_name = proj.name if proj else "Unknown Project"
                        mgr_name = proj.manager if proj else None
                        
                        admins = session.query(Admins).all()
                        for admin in admins:
                            new_notif = InAppNotification(
                                user_id=admin.id,
                                title="First Project Deployment",
                                message=f"Project '{proj_name}' has received its first deployment by {emp_name} (Task: {t_id}). Please verify the deployment.",
                                is_read=False
                            )
                            session.add(new_notif)
                
                # Check and update milestone status and actual_start date
                milestone_id = data.get("milestone_id")
                if milestone_id:
                    milestone = session.query(ProjectTimeline).filter_by(id=milestone_id).first()
                    if milestone:
                        if milestone.status == 'Pending':
                            milestone.status = 'Active'
                        if not milestone.actual_start:
                            milestone.actual_start = datetime.now()
                        
                # Update status of original handover task if handover_source_task_id is provided
                source_task_id = data.get("handover_source_task_id")
                if source_task_id:
                    orig_task = session.query(DeveloperTasks).filter_by(id=source_task_id).first()
                    if not orig_task:
                        orig_task = session.query(ContentCreatorTasks).filter_by(id=source_task_id).first()
                    if orig_task:
                        orig_task.task_status = data.get("task_status", "Completed")
                    
                    # Also update LeaveRequest if matches source_task_id
                    leave_req = session.query(LeaveRequest).filter_by(id=source_task_id).first()
                    if leave_req:
                        leave_req.handover_status = "Completed" if data.get("task_status") == "Completed" else "Pending"

                session.commit()
                session.refresh(new_task)


                logger.info(f"Content Task Added | EmpID: {employee_id} | Total Content: {data['total_content']}")
                return json.dumps(self.model_to_dict(new_task), indent=4, default=str)
            
            except ValueError as ve:
                session.rollback()
                return json.dumps({"error": "Invalid numerical data submitted for content counts.", "details": str(ve)})
            except Exception as e:
                session.rollback()
                return self._handle_error("add_content_creator_task", e, context=data)

    def get_all_tasks(self) -> str:
        """Utility to fetch all tasks from both tables, usually for high-level overviews."""
        with self.SessionLocal() as session:
            try:
                dev_tasks = session.query(DeveloperTasks).all()
                content_tasks = session.query(ContentCreatorTasks).all()
                
                combined = []
                for dt in dev_tasks:
                    d = self.model_to_dict(dt)
                    d['task_type'] = 'developer'
                    combined.append(d)
                for ct in content_tasks:
                    c = self.model_to_dict(ct)
                    c['task_type'] = 'content_creator'
                    combined.append(c)

                return json.dumps(combined, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_all_tasks", e)

    # ==========================================
    #     PROJECTS, SRS, & TIMELINE (PHASE 2)
    # ==========================================

    def add_project(self, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                # Sanitize incoming payload
                data = self._sanitize_payload(data)
                
                # Normalize spelling typo in referral field
                if "reffered_by" in data and "referred_by" not in data:
                    data["referred_by"] = data["reffered_by"]

                # Intercept and normalize legacy cost_type representing platform
                platforms = {"Mobile App", "Website", "Software", "Social Media", "Graphics"}
                if "cost_type" in data and data["cost_type"] in platforms:
                    if "project_platform" not in data or not data["project_platform"] or data["project_platform"] == "Generic project":
                        data["project_platform"] = data["cost_type"]
                    data["cost_type"] = "Internal / Non-Billable"

                valid_keys = {c.key for c in inspect(Projects).mapper.column_attrs}
                filtered_data = {k: v for k, v in data.items() if k in valid_keys}
                
                # Omit None values to allow Python-defined defaults to be used in model constructor
                filtered_data = {k: v for k, v in filtered_data.items() if v is not None}
                
                new_project = Projects(**filtered_data)
                session.add(new_project)
                session.commit()
                session.refresh(new_project)
                return json.dumps(self.model_to_dict(new_project), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("add_project", e, context=data)

    def _calculate_project_progress(self, session, proj) -> str:
        """
        System of Control Fix: Progress is now measured by Work Done (Milestones), 
        not by Money Spent (Budget).
        Falls back to manually set database progress if no milestones are defined.
        """
        try:
            total_milestones = session.query(ProjectTimeline).filter_by(project_id=proj.id).count()
            if total_milestones == 0:
                return proj.progress if hasattr(proj, 'progress') and proj.progress else "0%"
            
            completed_milestones = session.query(ProjectTimeline).filter(
                ProjectTimeline.project_id == proj.id,
                ProjectTimeline.status.ilike("%completed%")
            ).count()
            
            progress_pct = int((completed_milestones / total_milestones) * 100)
            return f"{progress_pct}%"
        except Exception:
            return proj.progress if hasattr(proj, 'progress') and proj.progress else "0%"

    def _calculate_payment_stats(self, session, proj) -> dict:
        """
        Computes the financial state of a project on the fly.
        """
        try:
            payments = session.query(ProjectPayments).filter_by(project_id=proj.id).all()
            total_paid = sum(p.amount for p in payments)
            client_cost = float(proj.client_cost or 0.0)
            pending_amount = max(0.0, client_cost - total_paid)
            
            if total_paid >= client_cost and client_cost > 0:
                payment_status = "Paid in Full"
            elif total_paid > 0:
                payment_status = "Partial"
            else:
                payment_status = "Unpaid"
                
            return {
                "total_paid": total_paid,
                "pending_amount": pending_amount,
                "payment_status": payment_status
            }
        except Exception:
            return {
                "total_paid": 0.0,
                "pending_amount": float(proj.client_cost or 0.0),
                "payment_status": "Unpaid"
            }

    def get_all_projects(self) -> str:
        with self.SessionLocal() as session:
            try:
                # Aggregate Payments
                payment_agg = session.query(
                    ProjectPayments.project_id,
                    func.sum(ProjectPayments.amount).label("total_paid")
                ).group_by(ProjectPayments.project_id).subquery()
                
                # Aggregate Timeline (Milestones)
                timeline_agg = session.query(
                    ProjectTimeline.project_id,
                    func.count(ProjectTimeline.id).label("total_milestones"),
                    func.sum(case((ProjectTimeline.status.ilike("%completed%"), 1), else_=0)).label("completed_milestones")
                ).group_by(ProjectTimeline.project_id).subquery()
                
                # Main Query with Joins
                query = session.query(
                    Projects,
                    func.coalesce(payment_agg.c.total_paid, 0).label("total_paid"),
                    func.coalesce(timeline_agg.c.total_milestones, 0).label("total_milestones"),
                    func.coalesce(timeline_agg.c.completed_milestones, 0).label("completed_milestones")
                ).outerjoin(
                    payment_agg, Projects.id == payment_agg.c.project_id
                ).outerjoin(
                    timeline_agg, Projects.id == timeline_agg.c.project_id
                )
                
                results = query.all()
                
                # Pre-calculate project accumulated costs
                dev_costs = {row.project_id: float(row.cost or 0) for row in session.query(
                    DeveloperTasks.project_id, func.sum(DeveloperTasks.employee_cost).label('cost')
                ).group_by(DeveloperTasks.project_id).all() if row.project_id}
                
                content_costs = {row.project_id: float(row.cost or 0) for row in session.query(
                    ContentCreatorTasks.project_id, func.sum(ContentCreatorTasks.employee_cost).label('cost')
                ).group_by(ContentCreatorTasks.project_id).all() if row.project_id}
                
                srs_project_ids = {row.project_id for row in session.query(SRS_Documents.project_id).all() if row.project_id}
                
                now = datetime.now()
                res_list = []
                for proj, total_paid, total_milestones, completed_milestones in results:
                    p_dict = self.model_to_dict(proj)
                    
                    # Progress calculation
                    p_dict['progress'] = proj.progress if proj.progress else "0%"
                        
                    # Payment calculation
                    client_cost = float(proj.client_cost or 0.0)
                    total_paid = float(total_paid or 0.0)
                    pending_amount = max(0.0, client_cost - total_paid)
                    
                    if total_paid >= client_cost and client_cost > 0:
                        payment_status = "Paid in Full"
                    elif total_paid > 0:
                        payment_status = "Partial"
                    else:
                        payment_status = "Unpaid"
                        
                    p_dict['total_paid'] = total_paid
                    p_dict['pending_amount'] = pending_amount
                    p_dict['payment_status'] = payment_status
                    
                    # Cost & Profit calculation
                    cost = dev_costs.get(proj.id, 0.0) + content_costs.get(proj.id, 0.0)
                    p_dict['accumulated_cost'] = cost
                    p_dict['profit'] = client_cost - cost
                    
                    # Risk status calculation
                    is_at_risk = False
                    risk_reasons = []
                    
                    p_status = (proj.status or "").lower()
                    if "risk" in p_status or "delayed" in p_status or "critical" in p_status:
                        is_at_risk = True
                        risk_reasons.append(f"Status is '{proj.status}'")
                        
                    p_budget = float(proj.budget) if proj.budget else 0.0
                    if p_budget > 0.0 and cost >= p_budget:
                        is_at_risk = True
                        risk_reasons.append(f"Budget Overrun (${cost:.2f} / ${p_budget:.2f})")
                    elif p_budget > 0.0 and cost >= (p_budget * 0.9):
                        is_at_risk = True
                        risk_reasons.append(f"Nearing Budget Limit ({cost/p_budget*100:.1f}%)")
                        
                    if proj.end_date and proj.end_date != "N/A":
                        try:
                            p_end = datetime.strptime(proj.end_date.strip(), "%Y-%m-%d")
                            if p_end < now:
                                is_at_risk = True
                                risk_reasons.append("Project Deadline Passed")
                        except:
                            pass
                            
                    if proj.id not in srs_project_ids:
                        is_at_risk = True
                        risk_reasons.append("SRS Document Missing")
                        
                    p_dict['is_at_risk'] = is_at_risk
                    p_dict['risk_reasons'] = risk_reasons
                    
                    res_list.append(p_dict)
                    
                return json.dumps(res_list, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_all_projects", e)

    def add_srs_document(self, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                new_srs = SRS_Documents(**data)
                session.add(new_srs)
                session.flush() # Ensure ID is generated
                
                # Link the active SRS to the Project
                project_id = data.get("project_id")
                if project_id:
                    project = session.query(Projects).filter_by(id=project_id).first()
                    if project:
                        project.srs_id = new_srs.id
                        
                session.commit()
                session.refresh(new_srs)
                return json.dumps(self.model_to_dict(new_srs), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("add_srs_document", e, context=data)

    def get_srs_by_project(self, project_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                documents = session.query(SRS_Documents).filter_by(project_id=project_id).order_by(SRS_Documents.created_at.desc()).all()
                return json.dumps([self.model_to_dict(d) for d in documents], indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_srs_by_project", e)

    def delete_srs_document(self, srs_id: str) -> str:
        import os
        with self.SessionLocal() as session:
            try:
                document = session.query(SRS_Documents).filter_by(id=srs_id).first()
                if not document:
                    return json.dumps({"error": "SRS document not found"})
                
                # If local file, delete it
                if document.file_url_or_path and not document.file_url_or_path.startswith("http"):
                    if os.path.exists(document.file_url_or_path):
                        os.remove(document.file_url_or_path)
                
                session.delete(document)
                session.commit()
                return json.dumps({"message": "SRS document deleted successfully"})
            except Exception as e:
                session.rollback()
                return self._handle_error("delete_srs_document", e, context={"srs_id": srs_id})


    def add_project_timeline(self, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                new_milestone = ProjectTimeline(**data)
                session.add(new_milestone)
                session.flush()

                project_id = new_milestone.project_id
                total = session.query(func.count(ProjectTimeline.id)).filter_by(project_id=project_id).scalar()
                if total and total > 0:
                    completed = session.query(func.count(ProjectTimeline.id)).filter(
                        ProjectTimeline.project_id == project_id,
                        ProjectTimeline.status.ilike('%completed%')
                    ).scalar()
                    proj = session.query(Projects).filter_by(id=project_id).first()
                    if proj:
                        proj.progress = f"{int((completed / total) * 100)}%"

                session.commit()
                session.refresh(new_milestone)
                return json.dumps(self.model_to_dict(new_milestone), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("add_project_timeline", e, context=data)

    def get_project_timeline(self, project_id: str, employee_id: str = None) -> str:
        with self.SessionLocal() as session:
            try:
                query = session.query(ProjectTimeline).filter_by(project_id=project_id)
                
                if employee_id:
                    # Filter milestones to show only those where the employee is assigned
                    query = query.join(
                        MilestoneAssignments, ProjectTimeline.id == MilestoneAssignments.milestone_id
                    ).filter(MilestoneAssignments.employee_id == employee_id)
                
                timeline = query.order_by(ProjectTimeline.expected_start).all()
                res_list = [self.model_to_dict(t) for t in timeline]
                return json.dumps(res_list, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_project_timeline", e, context={"project_id": project_id})

    def get_employee_assigned_milestones(self, employee_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                results = session.query(ProjectTimeline, Projects).join(
                    MilestoneAssignments, ProjectTimeline.id == MilestoneAssignments.milestone_id
                ).join(
                    Projects, ProjectTimeline.project_id == Projects.id
                ).filter(
                    MilestoneAssignments.employee_id == employee_id
                ).order_by(ProjectTimeline.expected_start).all()
                
                # Pre-calculate project costs
                dev_costs = {row.project_id: float(row.cost or 0) for row in session.query(
                    DeveloperTasks.project_id, func.sum(DeveloperTasks.employee_cost).label('cost')
                ).group_by(DeveloperTasks.project_id).all()}
                
                content_costs = {row.project_id: float(row.cost or 0) for row in session.query(
                    ContentCreatorTasks.project_id, func.sum(ContentCreatorTasks.employee_cost).label('cost')
                ).group_by(ContentCreatorTasks.project_id).all()}
                
                now = datetime.now()
                res_list = []
                for milestone, project in results:
                    m_dict = self.model_to_dict(milestone)
                    m_dict['projectName'] = project.name
                    
                    # Parent project risk
                    proj_cost = dev_costs.get(project.id, 0.0) + content_costs.get(project.id, 0.0)
                    proj_budget = float(project.budget) if project.budget else 0.0
                    proj_at_risk = False
                    proj_risk_reasons = []
                    
                    p_status = (project.status or "").lower()
                    if "risk" in p_status or "delayed" in p_status or "critical" in p_status:
                        proj_at_risk = True
                        proj_risk_reasons.append(f"Status: {project.status}")
                    if proj_budget > 0.0 and proj_cost >= proj_budget:
                        proj_at_risk = True
                        proj_risk_reasons.append("Budget Exceeded")
                    elif proj_budget > 0.0 and proj_cost >= (proj_budget * 0.9):
                        proj_at_risk = True
                        proj_risk_reasons.append("Nearing Budget Limit")
                    if project.end_date and project.end_date != "N/A":
                        try:
                            p_end = datetime.strptime(project.end_date, "%Y-%m-%d")
                            if p_end < now:
                                proj_at_risk = True
                                proj_risk_reasons.append("Deadline Passed")
                        except:
                            pass
                            
                    # Milestone risk
                    is_delayed = False
                    m_status = (milestone.status or "").lower()
                    if m_status == "delayed" or m_status == "at risk":
                        is_delayed = True
                    if milestone.expected_end and milestone.expected_end < now and m_status != "completed":
                        is_delayed = True
                        
                    m_dict['is_delayed'] = is_delayed
                    m_dict['project_at_risk'] = proj_at_risk
                    m_dict['project_risk_reasons'] = proj_risk_reasons
                    res_list.append(m_dict)
                return json.dumps(res_list, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_employee_assigned_milestones", e, context={"employee_id": employee_id})
                
    def assign_employee_to_milestone(self, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                existing = session.query(MilestoneAssignments).filter_by(
                    milestone_id=data.get("milestone_id"),
                    employee_id=data.get("employee_id")
                ).first()
                if existing:
                    return json.dumps({"error": "Employee is already assigned to this milestone."})
                
                new_assignment = MilestoneAssignments(**data)
                session.add(new_assignment)
                
                # Fetch milestone and project details to construct the notification message
                milestone = session.query(ProjectTimeline).filter_by(id=data.get("milestone_id")).first()
                if milestone:
                    proj = session.query(Projects).filter_by(id=milestone.project_id).first()
                    proj_name = proj.name if proj else "Unknown Project"
                    
                    new_notif = InAppNotification(
                        user_id=data.get("employee_id"),
                        title="New Milestone Assignment",
                        message=f"You have been assigned to milestone '{milestone.milestone_name}' in project '{proj_name}'.",
                        is_read=False
                    )
                    session.add(new_notif)

                session.commit()
                session.refresh(new_assignment)
                return json.dumps(self.model_to_dict(new_assignment), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("assign_employee_to_milestone", e, context=data)

    def get_milestone_assignments(self, milestone_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                assignments = session.query(MilestoneAssignments, Employees, Departments_Roles).join(
                    Employees, MilestoneAssignments.employee_id == Employees.id
                ).outerjoin(
                    Departments_Roles, Employees.role_id == Departments_Roles.id
                ).filter(MilestoneAssignments.milestone_id == milestone_id).all()
                
                res_list = []
                for assign, emp, role in assignments:
                    a_dict = self.model_to_dict(assign)
                    a_dict['full_name'] = emp.full_name
                    a_dict['job_title'] = role.role_name if role else "Employee"
                    a_dict['photo'] = emp.photo
                    res_list.append(a_dict)
                    
                return json.dumps(res_list, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_milestone_assignments", e, context={"milestone_id": milestone_id})

    def unassign_employee_from_milestone(self, assignment_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                assignment = session.query(MilestoneAssignments).filter_by(id=assignment_id).first()
                if not assignment:
                    return json.dumps({"error": "Milestone assignment not found."})
                session.delete(assignment)
                session.commit()
                return json.dumps({"message": "Employee unassigned from milestone successfully."})
            except Exception as e:
                session.rollback()
                return self._handle_error("unassign_employee_from_milestone", e, context={"assignment_id": assignment_id})

    # ==========================================
    #     PROJECT ASSIGNMENTS (PHASE 5)
    # ==========================================

    def assign_employee_to_project(self, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                # Check if already assigned
                existing = session.query(ProjectAssignments).filter_by(
                    project_id=data.get("project_id"),
                    employee_id=data.get("employee_id")
                ).first()
                if existing:
                    return json.dumps({"error": "Employee is already assigned to this project."})
                
                new_assignment = ProjectAssignments(**data)
                session.add(new_assignment)
                session.commit()
                session.refresh(new_assignment)
                return json.dumps(self.model_to_dict(new_assignment), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("assign_employee_to_project", e, context=data)

    def get_project_assignments(self, project_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                # Need to join with Employees and Departments_Roles to get names/rates/roles
                assignments = session.query(ProjectAssignments, Employees, Departments_Roles).join(
                    Employees, ProjectAssignments.employee_id == Employees.id
                ).outerjoin(
                    Departments_Roles, Employees.role_id == Departments_Roles.id
                ).filter(ProjectAssignments.project_id == project_id).all()
                
                res_list = []
                for assign, emp, role in assignments:
                    a_dict = self.model_to_dict(assign)
                    # Add employee details
                    a_dict['full_name'] = emp.full_name
                    a_dict['job_title'] = role.role_name if role else "Employee"
                    a_dict['hourly_cost_rate'] = emp.hourly_cost_rate
                    a_dict['hourly_billing_rate'] = emp.hourly_billing_rate
                    res_list.append(a_dict)
                    
                return json.dumps(res_list, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_project_assignments", e, context={"project_id": project_id})

    def unassign_employee(self, assignment_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                assignment = session.query(ProjectAssignments).filter_by(id=assignment_id).first()
                if not assignment:
                    return json.dumps({"error": "Assignment not found."})
                session.delete(assignment)
                session.commit()
                return json.dumps({"message": "Employee unassigned successfully."})
            except Exception as e:
                session.rollback()
                return self._handle_error("unassign_employee", e, context={"assignment_id": assignment_id})

    def edit_project_assignment(self, assignment_id: str, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                assignment = session.query(ProjectAssignments).filter_by(id=assignment_id).first()
                if not assignment:
                    return json.dumps({"error": "Assignment not found."})
                for key, value in data.items():
                    setattr(assignment, key, value)
                session.commit()
                session.refresh(assignment)
                return json.dumps(self.model_to_dict(assignment), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("edit_project_assignment", e, context=data)

    def get_employee_projects(self, employee_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                # Find all project assignments for this employee
                assignments = session.query(ProjectAssignments, Projects).join(
                    Projects, ProjectAssignments.project_id == Projects.id
                ).filter(
                    ProjectAssignments.employee_id == employee_id,
                    func.lower(func.coalesce(Projects.status, '')) != 'completed'
                ).all()
                
                # Pre-calculate project accumulated costs
                dev_costs = {row.project_id: float(row.cost or 0) for row in session.query(
                    DeveloperTasks.project_id, func.sum(DeveloperTasks.employee_cost).label('cost')
                ).group_by(DeveloperTasks.project_id).all()}
                
                content_costs = {row.project_id: float(row.cost or 0) for row in session.query(
                    ContentCreatorTasks.project_id, func.sum(ContentCreatorTasks.employee_cost).label('cost')
                ).group_by(ContentCreatorTasks.project_id).all()}
                
                now = datetime.now()
                res_list = []
                for assign, proj in assignments:
                    p_dict = self.model_to_dict(proj)
                    p_dict['progress'] = proj.progress if proj.progress else "0%"
                    
                    # Cost calculation
                    cost = dev_costs.get(proj.id, 0.0) + content_costs.get(proj.id, 0.0)
                    p_dict['accumulated_cost'] = cost
                    
                    # Risk status calculation
                    is_at_risk = False
                    risk_reasons = []
                    
                    p_status = (proj.status or "").lower()
                    if "risk" in p_status or "delayed" in p_status or "critical" in p_status:
                        is_at_risk = True
                        risk_reasons.append(f"Status is '{proj.status}'")
                        
                    p_budget = float(proj.budget) if proj.budget else 0.0
                    if p_budget > 0.0 and cost >= p_budget:
                        is_at_risk = True
                        risk_reasons.append(f"Budget Overrun (${cost:.2f} / ${p_budget:.2f})")
                    elif p_budget > 0.0 and cost >= (p_budget * 0.9):
                        is_at_risk = True
                        risk_reasons.append(f"Nearing Budget Limit ({cost/p_budget*100:.1f}%)")
                        
                    if proj.end_date and proj.end_date != "N/A":
                        try:
                            p_end = datetime.strptime(proj.end_date, "%Y-%m-%d")
                            if p_end < now:
                                is_at_risk = True
                                risk_reasons.append("Project Deadline Passed")
                        except:
                            pass
                            
                    # Check missing SRS
                    srs_exists = session.query(SRS_Documents).filter_by(project_id=proj.id).first()
                    if not srs_exists:
                        is_at_risk = True
                        risk_reasons.append("SRS Document Missing")
                        
                    p_dict['is_at_risk'] = is_at_risk
                    p_dict['risk_reasons'] = risk_reasons
                    res_list.append(p_dict)
                    
                return json.dumps(res_list, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_employee_projects", e, context={"employee_id": employee_id})

    # ==========================================
    #          AUDIT & COMPLIANCE
    # ==========================================

    def log_login_history(self, user_id: str, user_role: str, ip_address: str = "Unknown", user_agent: str = "Unknown") -> None:
        """Fire-and-forget logging function for tracking security access."""
        with self.SessionLocal() as session:
            try:
                log_entry = LoginHistory(
                    user_id=user_id,
                    user_role=user_role,
                    login_timestamp=datetime.utcnow(),  # Record login timestamps in UTC!
                    ip_address=ip_address,
                    user_agent=user_agent
                )
                session.add(log_entry)
                session.commit()
                logger.info(f"LOGIN SUCCESS: {user_role} ({user_id}) from {ip_address}")
            except Exception as e:
                session.rollback()
                # We do not return JSON here because this is called passively during login flows.
                logger.error(f"Failed to log login history for user {user_id}: {str(e)}", exc_info=True)

    def parse_user_agent(self, user_agent: str) -> tuple:
        """Parses user agent to return a tuple of (browser, device/OS) as human-readable strings."""
        if not user_agent or user_agent == "Unknown":
            return "Unknown", "Unknown"
        
        ua = user_agent.lower()
        
        # Determine OS / Device
        if "windows phone" in ua:
            device = "Windows Phone"
        elif "win" in ua:
            if "nt 10.0" in ua:
                device = "Windows 10/11"
            elif "nt 6.3" in ua:
                device = "Windows 8.1"
            elif "nt 6.2" in ua:
                device = "Windows 8"
            elif "nt 6.1" in ua:
                device = "Windows 7"
            else:
                device = "Windows"
        elif "android" in ua:
            device = "Android Device"
        elif "ipad" in ua:
            device = "iPad"
        elif "iphone" in ua:
            device = "iPhone"
        elif "mac" in ua:
            device = "macOS"
        elif "linux" in ua:
            device = "Linux"
        else:
            device = "Other Device"
            
        # Determine Browser
        if "edg" in ua:
            browser = "Microsoft Edge"
        elif "chrome" in ua or "crios" in ua:
            browser = "Google Chrome"
        elif "safari" in ua:
            browser = "Apple Safari"
        elif "firefox" in ua or "fxios" in ua:
            browser = "Mozilla Firefox"
        elif "opr" in ua or "opera" in ua:
            browser = "Opera"
        elif "msie" in ua or "trident" in ua:
            browser = "Internet Explorer"
        else:
            browser = "Other Browser"
            
        return browser, device

    # ==========================================
    #     REMAINING CRUD OPERATIONS (PHASE 2)
    # ==========================================

    def get_employee(self, employee_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                emp = session.query(Employees).filter_by(id=employee_id).first()
                if not emp:
                    return json.dumps({"error": "Employee not found"})
                res = self.model_to_dict(emp)
                res.pop('password', None)
                return json.dumps(res, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_employee", e)

    def delete_employee(self, employee_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                emp = session.query(Employees).filter_by(id=employee_id).first()
                if not emp:
                    return json.dumps({"error": "Employee not found"})

                # --- CLEANUP: Delete the employee's upload folder from disk ---
                upload_folder = os.path.join("data", "uploads", "employees", employee_id)
                if os.path.isdir(upload_folder):
                    shutil.rmtree(upload_folder)
                    logger.info(f"Deleted employee upload folder: {upload_folder}")

                session.delete(emp)
                session.commit()
                return json.dumps({"message": "Employee and all associated documents deleted successfully"})
            except Exception as e:
                session.rollback()
                return self._handle_error("delete_employee", e)

    def get_project(self, project_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                proj = session.query(Projects).filter_by(id=project_id).first()
                if not proj:
                    return json.dumps({"error": "Project not found"})
                
                p_dict = self.model_to_dict(proj)

                # Fetch Assignments
                assignments = session.query(ProjectAssignments).filter_by(project_id=project_id).all()
                assigned_employees = []
                for assign in assignments:
                    emp = session.query(Employees).filter_by(id=assign.employee_id).first()
                    if emp:
                        e_dict = self.model_to_dict(emp)
                        e_dict['custom_hourly_cost'] = assign.custom_hourly_cost
                        e_dict['custom_hourly_billing'] = assign.custom_hourly_billing
                        assigned_employees.append(e_dict)
                
                p_dict['assigned_employees'] = assigned_employees
                p_dict['progress'] = proj.progress if proj.progress else "0%"
                
                # Compute Payment Stats
                pay_stats = self._calculate_payment_stats(session, proj)
                p_dict.update(pay_stats)
                
                return json.dumps(p_dict, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_project", e)

    def export_project_csv_data(self, project_id: str) -> Optional[str]:
        import csv
        import io
        
        with self.SessionLocal() as session:
            try:
                # 1. Fetch core project details
                proj = session.query(Projects).filter_by(id=project_id).first()
                if not proj:
                    return None
                
                # Fetch related data
                # Assignments
                assignments = session.query(ProjectAssignments, Employees, Departments_Roles).join(
                    Employees, ProjectAssignments.employee_id == Employees.id
                ).outerjoin(
                    Departments_Roles, Employees.role_id == Departments_Roles.id
                ).filter(ProjectAssignments.project_id == project_id).all()
                
                # Timeline
                timeline = session.query(ProjectTimeline).filter_by(project_id=project_id).order_by(ProjectTimeline.expected_start).all()
                
                # Payments
                payments = session.query(ProjectPayments).filter_by(project_id=project_id).order_by(ProjectPayments.payment_date.desc()).all()
                
                # Expenses
                expenses = session.query(ProjectExpenses).filter_by(project_id=project_id).order_by(ProjectExpenses.expense_date.desc()).all()
                
                # Developer Tasks
                dev_tasks = session.query(DeveloperTasks, Employees).join(
                    Employees, DeveloperTasks.employee_id == Employees.id
                ).filter(DeveloperTasks.project_id == project_id).order_by(DeveloperTasks.date.desc()).all()
                
                # Content Creator Tasks
                content_tasks = session.query(ContentCreatorTasks, Employees).join(
                    Employees, ContentCreatorTasks.employee_id == Employees.id
                ).filter(ContentCreatorTasks.project_id == project_id).order_by(ContentCreatorTasks.date.desc()).all()
                
                # Checklist items
                checklists = session.query(ProjectChecklistState, ChecklistTemplate).join(
                    ChecklistTemplate, ProjectChecklistState.checklist_id == ChecklistTemplate.id
                ).filter(ProjectChecklistState.project_id == project_id).all()
                
                # SRS
                srs_docs = session.query(SRS_Documents).filter_by(project_id=project_id).order_by(SRS_Documents.created_at.desc()).all()
                
                # Setup CSV writer
                output = io.StringIO()
                writer = csv.writer(output, lineterminator='\n')
                
                # --- SECTION 1: PROJECT METADATA ---
                writer.writerow(["=== PROJECT METADATA ==="])
                writer.writerow([
                    "Project ID", "Project Name", "Project Type", "Project Platform", "Description", "Status",
                    "Client Cost", "Budget", "Approx Cost", "Cost Type", "Start Date",
                    "End Date", "Progress", "Manager", "Client Name", "Team", 
                    "Referred By", "Filled By", "Assigned To", "Created At", "Updated At"
                ])
                writer.writerow([
                    proj.id, proj.name, proj.project_type, proj.project_platform or "N/A", proj.description, proj.status,
                    proj.client_cost, proj.budget, proj.approx_cost, proj.cost_type, proj.start_date,
                    proj.end_date, proj.progress or "0%", proj.manager, proj.client, proj.team,
                    proj.referred_by, proj.filled_by, proj.assigned_to, proj.created_at, proj.updated_at
                ])
                writer.writerow([]) # Empty spacer
                
                # --- SECTION 2: ASSIGNED TEAM MEMBERS ---
                writer.writerow(["=== ASSIGNED TEAM MEMBERS ==="])
                writer.writerow([
                    "Employee ID", "Full Name", "Email", "Role/Job Title", 
                    "Standard Cost Rate", "Standard Billing Rate", "Custom Cost Override", "Custom Billing Override", "Assigned At"
                ])
                if assignments:
                    for assign, emp, role in assignments:
                        writer.writerow([
                            emp.id, emp.full_name, emp.email, role.role_name if role else "Employee",
                            emp.hourly_cost_rate, emp.hourly_billing_rate, assign.custom_hourly_cost, assign.custom_hourly_billing, assign.assigned_at
                        ])
                else:
                    writer.writerow(["No team members assigned."])
                writer.writerow([])
                
                # --- SECTION 3: PROJECT TIMELINE & MILESTONES ---
                writer.writerow(["=== PROJECT TIMELINE & MILESTONES ==="])
                writer.writerow([
                    "Milestone ID", "Milestone Name", "Expected Start", "Expected End", 
                    "Actual Start", "Actual End", "Status", "Remarks", "Created At"
                ])
                if timeline:
                    for t in timeline:
                        writer.writerow([
                            t.id, t.milestone_name, t.expected_start, t.expected_end,
                            t.actual_start, t.actual_end, t.status, t.remarks, t.created_at
                        ])
                else:
                    writer.writerow(["No timeline milestones defined."])
                writer.writerow([])
                
                # --- SECTION 4: CLIENT PAYMENT HISTORY ---
                writer.writerow(["=== CLIENT PAYMENT HISTORY ==="])
                writer.writerow([
                    "Payment ID", "Amount", "Payment Date", "Payment Method", "Reference Number", "Remarks", "Logged At"
                ])
                if payments:
                    for p in payments:
                        writer.writerow([
                            p.id, p.amount, p.payment_date, p.payment_method, p.reference_number, p.remarks, p.created_at
                        ])
                else:
                    writer.writerow(["No client payments logged."])
                writer.writerow([])
                
                # --- SECTION 5: PROJECT EXPENSES ---
                writer.writerow(["=== PROJECT EXPENSES ==="])
                writer.writerow([
                    "Expense ID", "Expense Name", "Amount", "Expense Date", "Description", "Logged At"
                ])
                if expenses:
                    for e in expenses:
                        writer.writerow([
                            e.id, e.expense_name, e.amount, e.expense_date, e.description, e.created_at
                        ])
                else:
                    writer.writerow(["No project expenses logged."])
                writer.writerow([])
                
                # --- SECTION 6: DEVELOPER TIMESHEETS ---
                writer.writerow(["=== DEVELOPER TIMESHEETS ==="])
                writer.writerow([
                    "Task ID", "Employee Name", "Date Logged", "Hours Logged", "Tech Stack Used", 
                    "GitHub Link", "PR Created?", "Branch Name", "Commit Count", "Repository", "Task Performed", "Tomorrow's Plan", "Task Status", "Work Type", 
                    "Sprint", "Module", "Feature", "Ticket ID", "No Project Reason", 
                    "Employee Cost", "Billing Amount", "Profit/Loss", "Logged At"
                ])
                if dev_tasks:
                    for dt, emp in dev_tasks:
                        writer.writerow([
                            dt.id, emp.full_name, dt.date, dt.hours_logged, dt.tech_stack,
                            dt.github_link, getattr(dt, 'github_pr_created', 'No'), getattr(dt, 'github_branch_name', 'N/A'), getattr(dt, 'github_commit_count', 0), getattr(dt, 'github_repo_name', 'N/A'), dt.task_performed, dt.tomorrow_plan, getattr(dt, 'task_status', 'Completed'), getattr(dt, 'work_type', 'Development'),
                            getattr(dt, 'sprint', 'N/A'), getattr(dt, 'module', 'N/A'), getattr(dt, 'feature', 'N/A'), getattr(dt, 'ticket_id', 'N/A'), getattr(dt, 'no_project_reason', 'N/A'),
                            dt.employee_cost, dt.billing_amount, dt.profit_loss, dt.created_at
                        ])
                else:
                    writer.writerow(["No developer tasks logged."])
                writer.writerow([])
                
                # --- SECTION 7: CONTENT CREATOR TIMESHEETS ---
                writer.writerow(["=== CONTENT CREATOR TIMESHEETS ==="])
                writer.writerow([
                    "Task ID", "Employee Name", "Date Logged", "Hours Logged", "Reels Count", "Long Video Count",
                    "Poster Count", "Calls Made", "Platform", "Total Content", "Task Performed", 
                    "Task Status", "Work Type", "Sprint", "Module", "Feature", "Ticket ID", "No Project Reason", 
                    "Employee Cost", "Billing Amount", "Profit/Loss", "Logged At"
                ])
                if content_tasks:
                    for ct, emp in content_tasks:
                        writer.writerow([
                            ct.id, emp.full_name, ct.date, ct.hours_logged, ct.reels_count, ct.long_video_count,
                            ct.poster_count, ct.calls_made, ct.platform, ct.total_content, ct.task_performed,
                            getattr(ct, 'task_status', 'Completed'), getattr(ct, 'work_type', 'Development'), getattr(ct, 'sprint', 'N/A'), getattr(ct, 'module', 'N/A'), getattr(ct, 'feature', 'N/A'), getattr(ct, 'ticket_id', 'N/A'), getattr(ct, 'no_project_reason', 'N/A'),
                            ct.employee_cost, ct.billing_amount, ct.profit_loss, ct.created_at
                        ])
                else:
                    writer.writerow(["No content creator tasks logged."])
                writer.writerow([])
                
                # --- SECTION 8: GATEKEEPER CHECKLIST ---
                writer.writerow(["=== GATEKEEPER CHECKLIST ==="])
                writer.writerow([
                    "Checklist Item ID", "Phase", "Task Description", "Completed State", "Last Updated"
                ])
                if checklists:
                    for state_item, template in checklists:
                        writer.writerow([
                            state_item.id, template.phase, template.task_description, 
                            "Checked" if state_item.is_checked else "Pending", state_item.updated_at
                        ])
                else:
                    writer.writerow(["No gatekeeper checklists initialized."])
                writer.writerow([])
                
                # --- SECTION 9: SRS & PROJECT DOCUMENTS ---
                writer.writerow(["=== SRS & PROJECT DOCUMENTS ==="])
                writer.writerow([
                    "Document ID", "Title", "Version", "File URL / Path", "Approval Status", "Approved By", "Uploaded At"
                ])
                if srs_docs:
                    for doc in srs_docs:
                        writer.writerow([
                            doc.id, doc.document_title, doc.version, doc.file_url_or_path, doc.status, doc.approved_by, doc.created_at
                        ])
                else:
                    writer.writerow(["No SRS documents linked."])
                
                return output.getvalue()
            except Exception as e:
                return self._handle_error("export_project_csv_data", e, context={"project_id": project_id})

    def export_project_xlsx_data(self, project_id: str) -> Optional[bytes]:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        from datetime import datetime, date

        def fmt_dt(val) -> str:
            if not val:
                return "N/A"
            if isinstance(val, str):
                return val
            try:
                return val.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                return str(val)

        def fmt_date(val) -> str:
            if not val:
                return "N/A"
            if isinstance(val, str):
                return val
            try:
                return val.strftime("%Y-%m-%d")
            except Exception:
                return str(val)
        
        with self.SessionLocal() as session:
            try:
                # 1. Fetch core project details
                proj = session.query(Projects).filter_by(id=project_id).first()
                if not proj:
                    return None
                
                # Fetch related data
                # Assignments
                assignments = session.query(ProjectAssignments, Employees, Departments_Roles).join(
                    Employees, ProjectAssignments.employee_id == Employees.id
                ).outerjoin(
                    Departments_Roles, Employees.role_id == Departments_Roles.id
                ).filter(ProjectAssignments.project_id == project_id).all()
                
                # Timeline
                timeline = session.query(ProjectTimeline).filter_by(project_id=project_id).order_by(ProjectTimeline.expected_start).all()
                
                # Payments
                payments = session.query(ProjectPayments).filter_by(project_id=project_id).order_by(ProjectPayments.payment_date.desc()).all()
                
                # Expenses
                expenses = session.query(ProjectExpenses).filter_by(project_id=project_id).order_by(ProjectExpenses.expense_date.desc()).all()
                
                # Developer Tasks
                dev_tasks = session.query(DeveloperTasks, Employees).join(
                    Employees, DeveloperTasks.employee_id == Employees.id
                ).filter(DeveloperTasks.project_id == project_id).order_by(DeveloperTasks.date.desc()).all()
                
                # Content Creator Tasks
                content_tasks = session.query(ContentCreatorTasks, Employees).join(
                    Employees, ContentCreatorTasks.employee_id == Employees.id
                ).filter(ContentCreatorTasks.project_id == project_id).order_by(ContentCreatorTasks.date.desc()).all()
                
                # Checklist items
                checklists = session.query(ProjectChecklistState, ChecklistTemplate).join(
                    ChecklistTemplate, ProjectChecklistState.checklist_id == ChecklistTemplate.id
                ).filter(ProjectChecklistState.project_id == project_id).all()
                
                # SRS
                srs_docs = session.query(SRS_Documents).filter_by(project_id=project_id).order_by(SRS_Documents.created_at.desc()).all()

                # Build openpyxl Workbook
                wb = Workbook()
                # Remove default sheet
                default_sheet = wb.active
                wb.remove(default_sheet)

                # Styles
                header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # brand indigo
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                data_font = Font(name="Calibri", size=11)
                italic_font = Font(name="Calibri", size=11, italic=True, color="7F7F7F")
                
                border_thin = Side(border_style="thin", color="D1D5DB")
                cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

                left_align = Alignment(horizontal="left", vertical="center")
                center_align = Alignment(horizontal="center", vertical="center")
                right_align = Alignment(horizontal="right", vertical="center")

                # Helper to write sheet
                def write_sheet(title: str, headers: List[str], rows: List[List[Any]], column_alignments: List[Alignment] = None, numeric_formats: Dict[int, str] = None):
                    ws = wb.create_sheet(title=title)
                    ws.views.sheetView[0].showGridLines = True
                    
                    # Write Headers
                    ws.append(headers)
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = cell_border
                    
                    # Write Data
                    if rows:
                        for r_idx, row in enumerate(rows, start=2):
                            ws.append(row)
                            for col_idx, val in enumerate(row, start=1):
                                cell = ws.cell(row=r_idx, column=col_idx)
                                cell.font = data_font
                                cell.border = cell_border
                                
                                # Alignments
                                if column_alignments and col_idx - 1 < len(column_alignments):
                                    cell.alignment = column_alignments[col_idx - 1]
                                else:
                                    cell.alignment = left_align
                                    
                                # Number Formats
                                if numeric_formats and (col_idx - 1) in numeric_formats:
                                    cell.number_format = numeric_formats[col_idx - 1]
                    else:
                        # Write standard placeholder
                        placeholder = ["No recorded database entries for this section."] + [""] * (len(headers) - 1)
                        ws.append(placeholder)
                        for col_idx in range(1, len(headers) + 1):
                            cell = ws.cell(row=2, column=col_idx)
                            cell.font = italic_font
                            cell.border = cell_border
                            cell.alignment = left_align
                    
                    # Adjust column widths
                    for col in ws.columns:
                        max_len = 0
                        for cell in col:
                            if cell.value:
                                val_str = str(cell.value)
                                if len(val_str) > max_len:
                                    max_len = len(val_str)
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

                # 1. Project Metadata
                proj_headers = [
                    "Project ID", "Project Name", "Project Type", "Project Platform", "Description", "Status",
                    "Client Cost", "Budget", "Approx Cost", "Cost Type", "Start Date",
                    "End Date", "Progress", "Manager", "Client Name", "Team", 
                    "Referred By", "Filled By", "Assigned To", "Created At", "Updated At"
                ]
                # Format Dates/Times
                start_date_str = fmt_date(proj.start_date)
                end_date_str = fmt_date(proj.end_date)
                created_at_str = fmt_dt(proj.created_at)
                updated_at_str = fmt_dt(proj.updated_at)
                
                proj_rows = [[
                    proj.id, proj.name, proj.project_type, proj.project_platform or "N/A", proj.description or "N/A", proj.status,
                    proj.client_cost or 0.0, proj.budget or 0.0, proj.approx_cost or 0.0, proj.cost_type or "N/A", start_date_str,
                    end_date_str, proj.progress or "0%", proj.manager or "N/A", proj.client or "N/A", proj.team or "N/A",
                    proj.referred_by or "N/A", proj.filled_by or "N/A", proj.assigned_to or "N/A", created_at_str, updated_at_str
                ]]
                
                # Alignments: ID(C), Name(L), Type(C), Platform(C), Desc(L), Status(C), ClientCost(R), Budget(R), ApproxCost(R), CostType(C), Dates(C), Progress(C), etc.
                proj_alignments = [
                    center_align, left_align, center_align, center_align, left_align, center_align,
                    right_align, right_align, right_align, center_align, center_align,
                    center_align, center_align, left_align, left_align, left_align,
                    left_align, left_align, left_align, center_align, center_align
                ]
                # Numeric Formats: Client Cost (index 6), Budget (index 7), Approx Cost (index 8)
                proj_formats = {6: "₹#,##0.00", 7: "₹#,##0.00", 8: "₹#,##0.00"}
                
                write_sheet("Project Metadata", proj_headers, proj_rows, proj_alignments, proj_formats)

                # 2. Assigned Team Members
                team_headers = [
                    "Employee ID", "Full Name", "Email", "Role/Job Title", 
                    "Standard Cost Rate", "Standard Billing Rate", "Custom Cost Override", "Custom Billing Override", "Assigned At"
                ]
                team_rows = []
                for assign, emp, role in assignments:
                    assigned_str = fmt_dt(assign.assigned_at)
                    team_rows.append([
                        emp.id, emp.full_name, emp.email, role.role_name if role else "Employee",
                        emp.hourly_cost_rate or 0.0, emp.hourly_billing_rate or 0.0, assign.custom_hourly_cost, assign.custom_hourly_billing, assigned_str
                    ])
                team_alignments = [
                    center_align, left_align, left_align, left_align,
                    right_align, right_align, right_align, right_align, center_align
                ]
                team_formats = {4: "₹#,##0.00", 5: "₹#,##0.00", 6: "₹#,##0.00", 7: "₹#,##0.00"}
                
                write_sheet("Assigned Team Members", team_headers, team_rows, team_alignments, team_formats)

                # 3. Timeline & Milestones
                timeline_headers = [
                    "Milestone ID", "Milestone Name", "Expected Start", "Expected End", 
                    "Actual Start", "Actual End", "Status", "Remarks", "Created At"
                ]
                timeline_rows = []
                for t in timeline:
                    exp_start = fmt_date(t.expected_start)
                    exp_end = fmt_date(t.expected_end)
                    act_start = fmt_date(t.actual_start)
                    act_end = fmt_date(t.actual_end)
                    created_at = fmt_dt(t.created_at)
                    timeline_rows.append([
                        t.id, t.milestone_name, exp_start, exp_end,
                        act_start, act_end, t.status, t.remarks or "N/A", created_at
                    ])
                timeline_alignments = [
                    center_align, left_align, center_align, center_align,
                    center_align, center_align, center_align, left_align, center_align
                ]
                
                write_sheet("Timeline & Milestones", timeline_headers, timeline_rows, timeline_alignments)

                # 4. Payments
                payment_headers = [
                    "Payment ID", "Amount", "Payment Date", "Payment Method", "Reference Number", "Remarks", "Logged At"
                ]
                payment_rows = []
                for p in payments:
                    pay_date = fmt_date(p.payment_date)
                    created_at = fmt_dt(p.created_at)
                    payment_rows.append([
                        p.id, p.amount or 0.0, pay_date, p.payment_method or "N/A", p.reference_number or "N/A", p.remarks or "N/A", created_at
                    ])
                payment_alignments = [
                    center_align, right_align, center_align, center_align, center_align, left_align, center_align
                ]
                payment_formats = {1: "₹#,##0.00"}
                
                write_sheet("Payments", payment_headers, payment_rows, payment_alignments, payment_formats)

                # 5. Expenses
                expense_headers = [
                    "Expense ID", "Expense Name", "Amount", "Expense Date", "Description", "Logged At"
                ]
                expense_rows = []
                for e in expenses:
                    exp_date = fmt_date(e.expense_date)
                    created_at = fmt_dt(e.created_at)
                    expense_rows.append([
                        e.id, e.expense_name, e.amount or 0.0, exp_date, e.description or "N/A", created_at
                    ])
                expense_alignments = [
                    center_align, left_align, right_align, center_align, left_align, center_align
                ]
                expense_formats = {2: "₹#,##0.00"}
                
                write_sheet("Expenses", expense_headers, expense_rows, expense_alignments, expense_formats)

                # 6. Developer Timesheets
                dev_headers = [
                    "Task ID", "Employee Name", "Date Logged", "Hours Logged", "Tech Stack Used", 
                    "GitHub Link", "PR Created?", "Branch Name", "Commit Count", "Repository", "Task Performed", "Tomorrow's Plan", "Task Status", "Work Type", 
                    "Sprint", "Module", "Feature", "Ticket ID", "No Project Reason", 
                    "Employee Cost", "Billing Amount", "Profit/Loss", "Logged At"
                ]
                dev_rows = []
                for dt, emp in dev_tasks:
                    dt_date = fmt_date(dt.date)
                    created_at = fmt_dt(dt.created_at)
                    dev_rows.append([
                        dt.id, emp.full_name, dt_date, dt.hours_logged or 0.0, dt.tech_stack or "N/A",
                        dt.github_link or "N/A", getattr(dt, 'github_pr_created', 'No'), getattr(dt, 'github_branch_name', 'N/A'), getattr(dt, 'github_commit_count', 0), getattr(dt, 'github_repo_name', 'N/A'), dt.task_performed or "N/A", dt.tomorrow_plan or "N/A",
                        getattr(dt, 'task_status', 'Completed'), getattr(dt, 'work_type', 'Development'),
                        getattr(dt, 'sprint', 'N/A'), getattr(dt, 'module', 'N/A'), getattr(dt, 'feature', 'N/A'), getattr(dt, 'ticket_id', 'N/A'), getattr(dt, 'no_project_reason', 'N/A'),
                        dt.employee_cost or 0.0, dt.billing_amount or 0.0, dt.profit_loss or 0.0, created_at
                    ])
                dev_alignments = [
                    center_align, left_align, center_align, right_align, center_align,
                    left_align, left_align, left_align, center_align, center_align,
                    center_align, center_align, center_align, center_align, left_align,
                    right_align, right_align, right_align, center_align
                ]
                dev_formats = {3: "0.0", 15: "₹#,##0.00", 16: "₹#,##0.00", 17: "₹#,##0.00"}
                
                write_sheet("Developer Timesheets", dev_headers, dev_rows, dev_alignments, dev_formats)

                # 7. Content Creator Timesheets
                content_headers = [
                    "Task ID", "Employee Name", "Date Logged", "Hours Logged", "Reels Count", "Long Video Count",
                    "Poster Count", "Calls Made", "Platform", "Total Content", "Task Performed", 
                    "Task Status", "Work Type", "Sprint", "Module", "Feature", "Ticket ID", "No Project Reason",
                    "Employee Cost", "Billing Amount", "Profit/Loss", "Logged At"
                ]
                content_rows = []
                for ct, emp in content_tasks:
                    ct_date = fmt_date(ct.date)
                    created_at = fmt_dt(ct.created_at)
                    content_rows.append([
                        ct.id, emp.full_name, ct_date, ct.hours_logged or 0.0, ct.reels_count or 0, ct.long_video_count or 0,
                        ct.poster_count or 0, ct.calls_made or 0, ct.platform or "N/A", ct.total_content or 0, ct.task_performed or "N/A",
                        getattr(ct, 'task_status', 'Completed'), getattr(ct, 'work_type', 'Development'),
                        getattr(ct, 'sprint', 'N/A'), getattr(ct, 'module', 'N/A'), getattr(ct, 'feature', 'N/A'), getattr(ct, 'ticket_id', 'N/A'), getattr(ct, 'no_project_reason', 'N/A'),
                        ct.employee_cost or 0.0, ct.billing_amount or 0.0, ct.profit_loss or 0.0, created_at
                    ])
                content_alignments = [
                    center_align, left_align, center_align, right_align, right_align, right_align,
                    right_align, right_align, center_align, right_align, left_align,
                    center_align, center_align, center_align, center_align, center_align, center_align, left_align,
                    right_align, right_align, right_align, center_align
                ]
                content_formats = {3: "0.0", 4: "#,##0", 5: "#,##0", 6: "#,##0", 7: "#,##0", 9: "#,##0", 18: "₹#,##0.00", 19: "₹#,##0.00", 20: "₹#,##0.00"}
                
                write_sheet("Content Creator Timesheets", content_headers, content_rows, content_alignments, content_formats)

                # 8. Gatekeeper Checklist
                checklist_headers = [
                    "Checklist Item ID", "Phase", "Task Description", "Completed State", "Last Updated"
                ]
                checklist_rows = []
                for state_item, template in checklists:
                    updated_at = fmt_dt(state_item.updated_at)
                    checklist_rows.append([
                        state_item.id, template.phase, template.task_description,
                        "Checked" if state_item.is_checked else "Pending", updated_at
                    ])
                checklist_alignments = [
                    center_align, center_align, left_align, center_align, center_align
                ]
                
                write_sheet("Gatekeeper Checklist", checklist_headers, checklist_rows, checklist_alignments)

                # 9. SRS & Documents
                srs_headers = [
                    "Document ID", "Title", "Version", "File URL / Path", "Approval Status", "Approved By", "Uploaded At"
                ]
                srs_rows = []
                for doc in srs_docs:
                    created_at = fmt_dt(doc.created_at)
                    srs_rows.append([
                        doc.id, doc.document_title, doc.version, doc.file_url_or_path or "N/A", doc.status, doc.approved_by or "N/A", created_at
                    ])
                srs_alignments = [
                    center_align, left_align, center_align, left_align, center_align, left_align, center_align
                ]
                
                write_sheet("SRS & Documents", srs_headers, srs_rows, srs_alignments)

                # Save Workbook to binary stream
                out = io.BytesIO()
                wb.save(out)
                return out.getvalue()
                
            except Exception as e:
                return self._handle_error("export_project_xlsx_data", e, context={"project_id": project_id})

    def _is_checklist_completed(self, session, project_id: str, phase: str) -> bool:
        templates = session.query(ChecklistTemplate).filter_by(project_id=project_id, phase=phase).all()
        if not templates:
            return True
        template_ids = [t.id for t in templates]
        checked_count = session.query(ProjectChecklistState).filter(
            ProjectChecklistState.project_id == project_id,
            ProjectChecklistState.checklist_id.in_(template_ids),
            ProjectChecklistState.is_checked == True
        ).count()
        return checked_count == len(templates)

    def edit_project(self, project_id: str, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                proj = session.query(Projects).filter_by(id=project_id).first()
                if not proj:
                    return json.dumps({"error": "Project not found"})
                
                # Sanitize incoming payload
                data = self._sanitize_payload(data)
                
                # Normalize spelling typo in referral field
                if "reffered_by" in data and "referred_by" not in data:
                    data["referred_by"] = data["reffered_by"]

                # Intercept and normalize legacy cost_type representing platform
                platforms = {"Mobile App", "Website", "Software", "Social Media", "Graphics"}
                if "cost_type" in data and data["cost_type"] in platforms:
                    if "project_platform" not in data or not data["project_platform"] or data["project_platform"] == "Generic project":
                        data["project_platform"] = data["cost_type"]
                    data["cost_type"] = "Internal / Non-Billable"

                # Check status transition rules
                if "status" in data:
                    new_status = data["status"]
                    if new_status == "In Progress" and proj.status != "In Progress":
                        if not self._is_checklist_completed(session, project_id, "START"):
                            return json.dumps({"error": "Cannot start project. The 'START' phase checklist must be fully completed first."})
                    elif new_status == "Completed" and proj.status != "Completed":
                        if not self._is_checklist_completed(session, project_id, "END"):
                            return json.dumps({"error": "Cannot complete project. The 'END' phase checklist must be fully completed first."})

                for key, value in data.items():
                    if hasattr(proj, key):
                        setattr(proj, key, value)
                session.commit()
                session.refresh(proj)
                return json.dumps(self.model_to_dict(proj), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("edit_project", e, context={"project_id": project_id})

    def delete_project(self, project_id: str) -> str:

        with self.SessionLocal() as session:
            try:
                proj = session.query(Projects).filter_by(id=project_id).first()
                if not proj:
                    return json.dumps({"error": "Project not found"})

                # --- CLEANUP: Delete all SRS files from disk ---
                srs_docs = session.query(SRS_Documents).filter_by(project_id=project_id).all()
                for doc in srs_docs:
                    if doc.file_url_or_path and not doc.file_url_or_path.startswith("http"):
                        if os.path.exists(doc.file_url_or_path):
                            os.remove(doc.file_url_or_path)
                            logger.info(f"Deleted SRS file: {doc.file_url_or_path}")

                # Delete the project's entire upload folder (data/uploads/projects/{project_id}_*)
                upload_pattern = os.path.join("data", "uploads", "projects", f"{project_id}_*")
                for folder_path in glob.glob(upload_pattern):
                    if os.path.isdir(folder_path):
                        shutil.rmtree(folder_path)
                        logger.info(f"Deleted SRS upload folder: {folder_path}")

                # Delete all payment records
                session.query(ProjectPayments).filter_by(project_id=project_id).delete()

                session.delete(proj)
                session.commit()
                return json.dumps({"message": "Project and all associated data deleted successfully"})
            except Exception as e:
                session.rollback()
                return self._handle_error("delete_project", e)

    def add_project_expense(self, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                expense = ProjectExpenses(**data)
                session.add(expense)
                session.commit()
                session.refresh(expense)
                return json.dumps(self.model_to_dict(expense), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("add_project_expense", e, context=data)

    def get_project_expenses(self, project_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                expenses = session.query(ProjectExpenses).filter_by(project_id=project_id).all()
                return json.dumps([self.model_to_dict(e) for e in expenses], indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_project_expenses", e)

    def delete_project_expense(self, expense_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                expense = session.query(ProjectExpenses).filter_by(id=expense_id).first()
                if not expense:
                    return json.dumps({"error": f"Expense not found."})
                session.delete(expense)
                session.commit()
                return json.dumps({"message": f"Expense deleted."})
            except Exception as e:
                session.rollback()
                return self._handle_error("delete_project_expense", e)

    def get_task(self, task_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                # Search Developer Tasks first
                task = session.query(DeveloperTasks).filter_by(id=task_id).first()
                if task:
                    res = self.model_to_dict(task)
                    res['task_type'] = 'developer'
                    return json.dumps(res, indent=4, default=str)
                
                # Search Content Tasks second
                task = session.query(ContentCreatorTasks).filter_by(id=task_id).first()
                if task:
                    res = self.model_to_dict(task)
                    res['task_type'] = 'content_creator'
                    return json.dumps(res, indent=4, default=str)
                
                return json.dumps({"error": "Task not found"})
            except Exception as e:
                return self._handle_error("get_task", e)

    def delete_task(self, task_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                # Find the task in either table
                task = session.query(DeveloperTasks).filter_by(id=task_id).first()
                if not task:
                    task = session.query(ContentCreatorTasks).filter_by(id=task_id).first()
                    
                if not task:
                    return json.dumps({"error": "Task not found"})
                    
                session.delete(task)
                session.commit()
                return json.dumps({"message": "Task deleted successfully"})
            except Exception as e:
                session.rollback()
                return self._handle_error("delete_task", e)

    def get_tasks_by_employee(self, employee_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                dev_tasks = session.query(DeveloperTasks).filter_by(employee_id=employee_id).all()
                content_tasks = session.query(ContentCreatorTasks).filter_by(employee_id=employee_id).all()
                
                combined = []
                for dt in dev_tasks:
                    d = self.model_to_dict(dt)
                    d['task_type'] = 'developer'
                    combined.append(d)
                for ct in content_tasks:
                    c = self.model_to_dict(ct)
                    c['task_type'] = 'content_creator'
                    combined.append(c)
                    
                return json.dumps(combined, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_tasks_by_employee", e)

    # ==========================================
    #     EXTENDED CRUD: ADMINS, ROLES, TIMELINES, SRS & TASKS
    # ==========================================

    def get_all_admins(self) -> str:
        with self.SessionLocal() as session:
            try:
                admins = session.query(Admins).all()
                results = []
                for a in admins:
                    a_dict = self.model_to_dict(a)
                    a_dict.pop('password', None)
                    results.append(a_dict)
                return json.dumps(results, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_all_admins", e)

    def get_admin(self, admin_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                admin = session.query(Admins).filter_by(id=admin_id).first()
                if not admin:
                    return json.dumps({"error": "Admin not found"})
                res = self.model_to_dict(admin)
                res.pop('password', None)
                return json.dumps(res, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_admin", e)

    def edit_admin(self, admin_id: str, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                admin = session.query(Admins).filter_by(id=admin_id).first()
                if not admin:
                    return json.dumps({"error": "Admin not found"})
                for key, value in data.items():
                    if key == "password":
                        value = get_password_hash(value)
                    setattr(admin, key, value)
                session.commit()
                session.refresh(admin)
                res = self.model_to_dict(admin)
                res.pop('password', None)
                return json.dumps(res, indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("edit_admin", e, context={"admin_id": admin_id})

    def delete_admin(self, admin_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                admin = session.query(Admins).filter_by(id=admin_id).first()
                if not admin:
                    return json.dumps({"error": "Admin not found"})
                session.delete(admin)
                session.commit()
                return json.dumps({"message": "Admin deleted successfully"})
            except Exception as e:
                session.rollback()
                return self._handle_error("delete_admin", e)

    def edit_department_role(self, role_id: str, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                role = session.query(Departments_Roles).filter_by(id=role_id).first()
                if not role:
                    return json.dumps({"error": "Role not found"})
                for key, value in data.items():
                    setattr(role, key, value)
                session.commit()
                session.refresh(role)
                return json.dumps(self.model_to_dict(role), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("edit_department_role", e, context={"role_id": role_id})

    def delete_department_role(self, role_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                role = session.query(Departments_Roles).filter_by(id=role_id).first()
                if not role:
                    return json.dumps({"error": "Role not found"})
                session.delete(role)
                session.commit()
                return json.dumps({"message": "Role deleted successfully"})
            except Exception as e:
                session.rollback()
                return self._handle_error("delete_department_role", e)

    def edit_srs_document(self, srs_id: str, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                srs = session.query(SRS_Documents).filter_by(id=srs_id).first()
                if not srs:
                    return json.dumps({"error": "SRS not found"})
                for key, value in data.items():
                    setattr(srs, key, value)
                session.commit()
                session.refresh(srs)
                return json.dumps(self.model_to_dict(srs), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("edit_srs_document", e, context={"srs_id": srs_id})

    def edit_project_timeline(self, timeline_id: str, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                timeline = session.query(ProjectTimeline).filter_by(id=timeline_id).first()
                if not timeline:
                    return json.dumps({"error": "Timeline milestone not found"})
                for key, value in data.items():
                    setattr(timeline, key, value)
                
                session.flush()
                # Auto-update project progress
                project_id = timeline.project_id
                total = session.query(func.count(ProjectTimeline.id)).filter_by(project_id=project_id).scalar()
                if total and total > 0:
                    completed = session.query(func.count(ProjectTimeline.id)).filter(
                        ProjectTimeline.project_id == project_id,
                        ProjectTimeline.status.ilike('%completed%')
                    ).scalar()
                    proj = session.query(Projects).filter_by(id=project_id).first()
                    if proj:
                        proj.progress = f"{int((completed / total) * 100)}%"

                session.commit()
                session.refresh(timeline)
                return json.dumps(self.model_to_dict(timeline), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("edit_project_timeline", e, context={"timeline_id": timeline_id})

    def delete_project_timeline(self, timeline_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                timeline = session.query(ProjectTimeline).filter_by(id=timeline_id).first()
                if not timeline:
                    return json.dumps({"error": "Timeline milestone not found"})
                
                project_id = timeline.project_id
                session.delete(timeline)
                session.flush()
                
                # Auto-update project progress
                total = session.query(func.count(ProjectTimeline.id)).filter_by(project_id=project_id).scalar()
                proj = session.query(Projects).filter_by(id=project_id).first()
                if proj:
                    if total and total > 0:
                        completed = session.query(func.count(ProjectTimeline.id)).filter(
                            ProjectTimeline.project_id == project_id,
                            ProjectTimeline.status.ilike('%completed%')
                        ).scalar()
                        proj.progress = f"{int((completed / total) * 100)}%"
                    else:
                        proj.progress = "0%"

                session.commit()
                return json.dumps({"message": "Milestone deleted successfully"})
            except Exception as e:
                session.rollback()
                return self._handle_error("delete_project_timeline", e)

    def edit_developer_task(self, task_id: str, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                task = session.query(DeveloperTasks).filter_by(id=task_id).first()
                if not task:
                    return json.dumps({"error": "Developer Task not found."})
                
                for key, value in data.items():
                    setattr(task, key, value)
                
                # RE-CALCULATE FINANCIALS IF HOURS ARE EDITED
                if 'hours_logged' in data:
                    # check 24 hrs daily limit
                    task_cal_date = task.date.date()
                    start_of_day = datetime.combine(task_cal_date, datetime.min.time())
                    end_of_day = datetime.combine(task_cal_date, datetime.max.time())
                    
                    dev_hours_today = session.query(func.sum(DeveloperTasks.hours_logged)).filter(
                        DeveloperTasks.employee_id == task.employee_id,
                        DeveloperTasks.id != task.id,
                        DeveloperTasks.date.between(start_of_day, end_of_day)
                    ).scalar() or 0.0
                    
                    content_hours_today = session.query(func.sum(ContentCreatorTasks.hours_logged)).filter(
                        ContentCreatorTasks.employee_id == task.employee_id,
                        ContentCreatorTasks.date.between(start_of_day, end_of_day)
                    ).scalar() or 0.0
                    
                    existing_hours_today = float(dev_hours_today) + float(content_hours_today)
                    new_hours = float(data['hours_logged'])
                    if existing_hours_today + new_hours > 24.0:
                        return json.dumps({"error": f"Daily limit exceeded: You have already logged {existing_hours_today} hours on this date. Setting this task to {new_hours} hours would exceed the 24 hours daily limit."})

                    employee = session.query(Employees).filter_by(id=task.employee_id).first()
                    if employee:
                        cost_rate = float(employee.hourly_cost_rate or 0.0)
                        billing_rate = float(employee.hourly_billing_rate or 0.0)
                        task.employee_cost = float(task.hours_logged) * cost_rate
                        task.billing_amount = float(task.hours_logged) * billing_rate
                        task.profit_loss = task.billing_amount - task.employee_cost

                session.commit()
                session.refresh(task)
                res = self.model_to_dict(task)
                res['task_type'] = 'developer'
                return json.dumps(res, indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("edit_developer_task", e, context={"task_id": task_id})

    def edit_content_creator_task(self, task_id: str, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                task = session.query(ContentCreatorTasks).filter_by(id=task_id).first()
                if not task:
                    return json.dumps({"error": "Content Task not found."})
                
                for key, value in data.items():
                    setattr(task, key, value)
                
                # RE-CALCULATE FINANCIALS IF HOURS ARE EDITED
                if 'hours_logged' in data:
                    # check 24 hrs daily limit
                    task_cal_date = task.date.date()
                    start_of_day = datetime.combine(task_cal_date, datetime.min.time())
                    end_of_day = datetime.combine(task_cal_date, datetime.max.time())
                    
                    dev_hours_today = session.query(func.sum(DeveloperTasks.hours_logged)).filter(
                        DeveloperTasks.employee_id == task.employee_id,
                        DeveloperTasks.date.between(start_of_day, end_of_day)
                    ).scalar() or 0.0
                    
                    content_hours_today = session.query(func.sum(ContentCreatorTasks.hours_logged)).filter(
                        ContentCreatorTasks.employee_id == task.employee_id,
                        ContentCreatorTasks.id != task.id,
                        ContentCreatorTasks.date.between(start_of_day, end_of_day)
                    ).scalar() or 0.0
                    
                    existing_hours_today = float(dev_hours_today) + float(content_hours_today)
                    new_hours = float(data['hours_logged'])
                    if existing_hours_today + new_hours > 24.0:
                        return json.dumps({"error": f"Daily limit exceeded: You have already logged {existing_hours_today} hours on this date. Setting this task to {new_hours} hours would exceed the 24 hours daily limit."})

                    employee = session.query(Employees).filter_by(id=task.employee_id).first()
                    if employee:
                        cost_rate = float(employee.hourly_cost_rate or 0.0)
                        billing_rate = float(employee.hourly_billing_rate or 0.0)
                        task.employee_cost = float(task.hours_logged) * cost_rate
                        task.billing_amount = float(task.hours_logged) * billing_rate
                        task.profit_loss = task.billing_amount - task.employee_cost

                # RE-CALCULATE TOTAL CONTENT IF COUNTS ARE EDITED
                if any(k in data for k in ['reels_count', 'long_video_count', 'poster_count']):
                    task.total_content = (task.reels_count or 0) + (task.long_video_count or 0) + (task.poster_count or 0)

                session.commit()
                session.refresh(task)
                res = self.model_to_dict(task)
                res['task_type'] = 'content_creator'
                return json.dumps(res, indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("edit_content_creator_task", e, context={"task_id": task_id})

    # ==========================================
    #             ATTENDANCE & LEAVES
    # ==========================================
    def _auto_checkout_old_sessions_internal(self, session, employee_id: str, now: datetime) -> None:
        """
        Scans for any open attendance records (check_in_time present, check_out_time missing)
        for this employee that are older than 16 hours, and auto-completes them (8 hours duration)
        so that data remains clean and doesn't conflict with today's logins/check-ins.
        Also scans for overnight sessions spanning into today past the 10-minute shift reset boundary.
        """
        from datetime import timedelta
        employee = session.query(Employees).filter_by(id=employee_id).first()
        if not employee:
            return
        
        shift_start_str = employee.shift_start_time or "09:00"
        try:
            sh, sm = map(int, shift_start_str.split(":"))
        except Exception:
            sh, sm = 9, 0
            
        today_shift_reset = datetime.combine(now.date(), datetime.min.time().replace(hour=sh, minute=sm)) - timedelta(minutes=10)
        
        dangling_records = session.query(Attendance).filter(
            Attendance.employee_id == employee_id,
            Attendance.check_in_time.isnot(None),
            Attendance.check_out_time.is_(None)
        ).all()
        
        for r in dangling_records:
            if now - r.check_in_time > timedelta(hours=16):
                r.check_out_time = r.check_in_time + timedelta(hours=8)
                r.total_hours = 8.0
                r.notes = (r.notes + " | " if r.notes else "") + "Auto Checked Out (forgot to check out)"
                r.status = "Checked Out"
            elif r.check_in_time < today_shift_reset and now >= today_shift_reset:
                r.check_out_time = today_shift_reset
                time_diff = today_shift_reset - r.check_in_time
                r.total_hours = max(0.0, time_diff.total_seconds() / 3600.0)
                r.notes = (r.notes + " | " if r.notes else "") + "Auto Checked Out at shift reset (overnight shift)"
                r.status = "Checked Out"
        session.commit()

    def _auto_checkout_all_old_sessions_internal(self, session, now: datetime) -> None:
        """
        Scans for all unclosed attendance records across all employees and auto-checks them out
        if they are older than 16 hours or span past the employee's shift reset boundary.
        """
        from datetime import timedelta
        dangling_records = session.query(Attendance).filter(
            Attendance.check_in_time.isnot(None),
            Attendance.check_out_time.is_(None)
        ).all()
        
        for r in dangling_records:
            emp = session.query(Employees).filter_by(id=r.employee_id).first()
            if not emp:
                continue
            shift_start_str = emp.shift_start_time or "09:00"
            try:
                sh, sm = map(int, shift_start_str.split(":"))
            except Exception:
                sh, sm = 9, 0
                
            today_shift_reset = datetime.combine(now.date(), datetime.min.time().replace(hour=sh, minute=sm)) - timedelta(minutes=10)
            
            if now - r.check_in_time > timedelta(hours=16):
                r.check_out_time = r.check_in_time + timedelta(hours=8)
                r.total_hours = 8.0
                r.notes = (r.notes + " | " if r.notes else "") + "Auto Checked Out (forgot to check out)"
                r.status = "Checked Out"
            elif r.check_in_time < today_shift_reset and now >= today_shift_reset:
                r.check_out_time = today_shift_reset
                time_diff = today_shift_reset - r.check_in_time
                r.total_hours = max(0.0, time_diff.total_seconds() / 3600.0)
                r.notes = (r.notes + " | " if r.notes else "") + "Auto Checked Out at shift reset (overnight shift)"
                r.status = "Checked Out"
        session.commit()

    def check_in(self, employee_id: str, ip_address: str = None) -> str:
        with self.SessionLocal() as session:
            try:
                now = datetime.now()
                # Run auto-checkout first for old dangling sessions of this employee
                self._auto_checkout_old_sessions_internal(session, employee_id, now)

                # Check if there is an active session (not checked out)
                active_session = session.query(Attendance).filter(
                    Attendance.employee_id == employee_id,
                    Attendance.check_in_time.isnot(None),
                    Attendance.check_out_time.is_(None)
                ).first()
                
                if active_session:
                    return json.dumps({"error": "Already checked in. Please check out of your current session first."})

                employee = session.query(Employees).filter_by(id=employee_id).first()
                if not employee:
                    return json.dumps({"error": "Employee not found."})
                employee_name = employee.full_name

                # Evaluate timing rule based status
                shift_start_str = employee.shift_start_time or "09:00"
                shift_end_str = employee.shift_end_time or "18:00"
                try:
                    sh, sm = map(int, shift_start_str.split(":"))
                except Exception:
                    sh, sm = 9, 0
                try:
                    eh, em = map(int, shift_end_str.split(":"))
                except Exception:
                    eh, em = 18, 0

                from datetime import timedelta
                today_shift_start = datetime.combine(now.date(), datetime.min.time().replace(hour=sh, minute=sm))
                today_shift_end = datetime.combine(now.date(), datetime.min.time().replace(hour=eh, minute=em))
                today_shift_reset = today_shift_start - timedelta(minutes=10)

                if now < today_shift_reset:
                    target_date = now.date() - timedelta(days=1)
                    attendance_status = "Extended Shift"
                else:
                    target_date = now.date()
                    if now > today_shift_end:
                        attendance_status = "Extended Shift"
                    else:
                        diff = now - today_shift_start
                        diff_seconds = diff.total_seconds()
                        if diff_seconds > 4 * 3600:
                            attendance_status = "Absent"
                        elif diff_seconds > 2 * 3600:
                            attendance_status = "Half-Day"
                        elif diff_seconds > 15 * 60:
                            attendance_status = "Late"
                        else:
                            attendance_status = "Present"

                # If there's an Absent or On Leave record on the target date, we override it
                start_of_target_day = datetime.combine(target_date, datetime.min.time())
                end_of_target_day = datetime.combine(target_date, datetime.max.time())
                existing_record = session.query(Attendance).filter(
                    Attendance.employee_id == employee_id,
                    Attendance.date.between(start_of_target_day, end_of_target_day)
                ).first()
                
                if existing_record and existing_record.status in ["Absent", "On Leave"]:
                    existing_record.check_in_time = now
                    existing_record.status = "Checked In"
                    existing_record.attendance_status = attendance_status
                    existing_record.ip_address = ip_address
                    existing_record.notes = (existing_record.notes + " | " if existing_record.notes else "") + f"Checked in (overrode {existing_record.status}) | Attendance: {attendance_status}"
                    session.commit()
                    session.refresh(existing_record)
                    attendence_dict = self.model_to_dict(existing_record)
                    attendence_dict['employee_name'] = employee_name
                    return json.dumps(attendence_dict, indent=4, default=str)
                
                new_attendance = Attendance(
                    employee_id=employee_id,
                    check_in_time=now,
                    date=datetime.combine(target_date, now.time()),
                    status="Checked In",
                    attendance_status=attendance_status,
                    ip_address=ip_address,
                    notes=f"Checked in | Attendance: {attendance_status}"
                )

                session.add(new_attendance)
                session.commit()
                session.refresh(new_attendance)
                attendence_dict = self.model_to_dict(new_attendance)
                attendence_dict['employee_name'] = employee_name
                return json.dumps(attendence_dict, indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("check_in", e)

    def check_out(self, employee_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                now = datetime.now()
                # Run auto-checkout first for old dangling sessions of this employee
                self._auto_checkout_old_sessions_internal(session, employee_id, now)

                # Find the most recent active session (no check-out time)
                attendance = session.query(Attendance).filter(
                    Attendance.employee_id == employee_id,
                    Attendance.check_in_time.isnot(None),
                    Attendance.check_out_time.is_(None)
                ).order_by(Attendance.check_in_time.desc()).first()
                
                if not attendance:
                    return json.dumps({"error": "No active check-in record found. Please check in first."})
                
                attendance.check_out_time = now
                attendance.status = "Checked Out"
                # Calculate total hours
                time_diff = attendance.check_out_time - attendance.check_in_time
                attendance.total_hours = time_diff.total_seconds() / 3600.0

                session.commit()
                session.refresh(attendance)
                return json.dumps(self.model_to_dict(attendance), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("check_out", e)

    def record_daily_absences_internal(self, session) -> None:
        """Helper to dynamically populate 'Absent' or 'On Leave' records for active employees who haven't checked in today."""
        now = datetime.now()
        start_of_day = datetime.combine(now.date(), datetime.min.time())
        end_of_day = datetime.combine(now.date(), datetime.max.time())
        
        from src.database.database_tables import LeaveRequest
        today_str = now.strftime("%Y-%m-%d")

        active_employees = session.query(Employees).filter(Employees.is_active == True).all()
        for emp in active_employees:
            # Check if there is any attendance record for this employee today
            existing = session.query(Attendance).filter(
                Attendance.employee_id == emp.id,
                Attendance.date.between(start_of_day, end_of_day)
            ).first()
            
            if not existing:
                # Check if there is an approved leave request for today
                approved_leave = session.query(LeaveRequest).filter(
                    LeaveRequest.employee_id == emp.id,
                    LeaveRequest.status == "Approved",
                    LeaveRequest.start_date <= today_str,
                    LeaveRequest.end_date >= today_str
                ).first()

                status = "Absent"
                notes = "Auto-recorded absent (no check-in detected)"
                if approved_leave:
                    status = "On Leave"
                    notes = f"On Leave (Approved request: {approved_leave.reason})"

                absent_record = Attendance(
                    employee_id=emp.id,
                    date=now,
                    status=status,
                    attendance_status=status,
                    total_hours=0.0,
                    notes=notes
                )
                session.add(absent_record)
        session.commit()

    def record_daily_absences(self) -> str:
        with self.SessionLocal() as session:
            try:
                self.record_daily_absences_internal(session)
                return json.dumps({"status": "success", "message": "Daily absences recorded successfully."})
            except Exception as e:
                session.rollback()
                return self._handle_error("record_daily_absences", e)

    def get_all_attendance(self) -> str:
        with self.SessionLocal() as session:
            try:
                now = datetime.now()
                # Run auto-checkout for all dangling records first
                self._auto_checkout_all_old_sessions_internal(session, now)
                
                # Automatically record absences first!
                self.record_daily_absences_internal(session)
                
                # Fetch all attendance records joined with Employees to resolve names
                records = session.query(Attendance, Employees.full_name)\
                                 .join(Employees, Attendance.employee_id == Employees.id)\
                                 .order_by(Attendance.date.desc())\
                                 .all()
                
                results = []
                for r, name in records:
                    d = self.model_to_dict(r)
                    d['employee_name'] = name
                    results.append(d)
                return json.dumps(results, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_all_attendance", e)

    # ==========================================
    #          PROJECT PAYMENTS
    # ==========================================

    def add_project_payment(self, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                new_payment = ProjectPayments(**data)
                session.add(new_payment)
                session.commit()
                session.refresh(new_payment)
                return json.dumps(self.model_to_dict(new_payment), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("add_project_payment", e, context=data)

    def get_project_payments(self, project_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                payments = session.query(ProjectPayments).filter_by(project_id=project_id).order_by(ProjectPayments.payment_date.desc()).all()
                return json.dumps([self.model_to_dict(p) for p in payments], indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_project_payments", e, context={"project_id": project_id})

    def delete_project_payment(self, payment_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                payment = session.query(ProjectPayments).filter_by(id=payment_id).first()
                if not payment:
                    return json.dumps({"error": "Payment record not found."})
                session.delete(payment)
                session.commit()
                return json.dumps({"message": "Payment record deleted successfully."})
            except Exception as e:
                session.rollback()
                return self._handle_error("delete_project_payment", e, context={"payment_id": payment_id})

    def get_employee_attendance(self, employee_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                # Run auto-checkout first to ensure their displayed attendance history is clean
                self._auto_checkout_old_sessions_internal(session, employee_id, datetime.now())
                
                records = session.query(Attendance).filter_by(employee_id=employee_id).order_by(Attendance.date.desc()).all()
                return json.dumps([self.model_to_dict(r) for r in records], indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_employee_attendance", e)

    def get_all_login_history(self) -> str:
        with self.SessionLocal() as session:
            try:
                records = session.query(LoginHistory).order_by(LoginHistory.login_timestamp.desc()).all()
                
                # Fetch all employees and admins to build a fast map for metadata resolution
                employees = session.query(Employees.id, Employees.full_name, Employees.username).all()
                admins = session.query(Admins.id, Admins.full_name, Admins.username).all()
                
                user_map = {}
                for e_id, name, uname in employees:
                    user_map[e_id] = {"full_name": name, "username": uname}
                for a_id, name, uname in admins:
                    user_map[a_id] = {"full_name": name, "username": uname}
                
                results = []
                for r in records:
                    r_dict = self.model_to_dict(r)
                    
                    # Parse user agent into device/browser
                    ua = r.user_agent or "Unknown"
                    browser, device = self.parse_user_agent(ua)
                    
                    # Resolve names
                    user_info = user_map.get(r.user_id, {"full_name": "Unknown", "username": "N/A"})
                    r_dict["full_name"] = user_info["full_name"]
                    r_dict["username"] = user_info["username"]
                    r_dict["browser"] = browser
                    r_dict["device"] = device
                    
                    # Format login_timestamp as ISO-8601 string ending with Z (UTC indicator)
                    if isinstance(r.login_timestamp, datetime):
                        r_dict["login_timestamp"] = r.login_timestamp.strftime("%Y-%m-%dT%H:%M:%SZ")
                    
                    results.append(r_dict)
                    
                return json.dumps(results, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_all_login_history", e)

    def get_all_leave_requests(self) -> str:
        with self.SessionLocal() as session:
            try:
                records = session.query(LeaveRequest).all()
                return json.dumps([self.model_to_dict(r) for r in records], indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_all_leave_requests", e)

    def create_leave_request(self, employee_id: str, start_date: str, end_date: str, reason: str,
                             leave_type: str = "Paid Leave", half_day_option: str = "Full Day",
                             total_days: float = 1.0, pending_work_summary: str = "N/A",
                             backup_employee_id: str = "N/A", deployment_pending: str = "No",
                             project_id: str = None, milestone_id: str = None, task_type: str = None) -> str:
        with self.SessionLocal() as session:
            try:
                # 11. Add Mandatory Validation Rules:
                # Reason Field: Min 15 chars, no empty spaces only.
                stripped_reason = reason.strip()
                if not stripped_reason:
                    return json.dumps({"error": "Reason cannot be empty or contain only spaces."})
                if len(stripped_reason) < 15:
                    return json.dumps({"error": "Reason must be at least 15 characters long."})
                if len(stripped_reason) > 500:
                    return json.dumps({"error": "Reason cannot exceed 500 characters."})

                # Date parsing
                start_dt = self._parse_datetime(start_date) if isinstance(start_date, str) else start_date
                end_dt = self._parse_datetime(end_date) if isinstance(end_date, str) else end_date
                if not start_dt or not end_dt:
                    return json.dumps({"error": "Invalid date format."})

                # End date >= Start date
                if end_dt.date() < start_dt.date():
                    return json.dumps({"error": "End date cannot be before start date."})

                # Cannot select past locked dates (start date cannot be before today's date)
                if start_dt.date() < datetime.now().date():
                    return json.dumps({"error": "Cannot apply for leave starting in the past."})

                # Overlapping leave requests
                overlapping = session.query(LeaveRequest).filter(
                    LeaveRequest.employee_id == employee_id,
                    LeaveRequest.status.in_(["Pending", "Approved", "Cancellation Requested"]),
                    LeaveRequest.start_date <= end_dt,
                    LeaveRequest.end_date >= start_dt
                ).first()
                if overlapping:
                    if overlapping.status == "Approved":
                        return json.dumps({"error": "You already have an approved leave request for overlapping dates."})
                    else:
                        return json.dumps({"error": "You already have a pending leave request for overlapping dates."})

                new_request = LeaveRequest(
                    employee_id=employee_id,
                    start_date=start_dt,
                    end_date=end_dt,
                    reason=reason,
                    leave_type=leave_type,
                    half_day_option=half_day_option,
                    total_days=total_days,
                    pending_work_summary=pending_work_summary,
                    backup_employee_id=backup_employee_id,
                    deployment_pending=deployment_pending,
                    project_id=project_id,
                    milestone_id=milestone_id,
                    task_type=task_type,
                    handover_status="Pending",
                    status="Pending"
                )
                session.add(new_request)
                session.commit()
                session.refresh(new_request)

                # Write Audit Log
                self.write_audit_log(
                    user_id=employee_id,
                    action="SUBMIT_LEAVE_REQUEST",
                    target_id=new_request.id,
                    details={"start_date": str(start_date), "end_date": str(end_date), "leave_type": leave_type}
                )

                # 10. Add Notification: Manager on new leave request
                admins = session.query(Admins).all()
                emp = session.query(Employees).filter_by(id=employee_id).first()
                emp_name = emp.full_name if emp else "Employee"
                for admin in admins:
                    new_notif = InAppNotification(
                        user_id=admin.id,
                        title="New Leave Request",
                        message=f"{emp_name} has applied for {leave_type} ({total_days} days) from {start_dt.strftime('%Y-%m-%d')} to {end_dt.strftime('%Y-%m-%d')}.",
                        is_read=False
                    )
                    session.add(new_notif)

                # Commented out to prevent notifying the backup employee before admin approval
                # if backup_employee_id and backup_employee_id != "N/A":
                #     backup_notif = InAppNotification(
                #         user_id=backup_employee_id,
                #         title="Task Handover Assigned",
                #         message=f"{emp_name} has designated you as backup for their leave from {start_dt.strftime('%Y-%m-%d')} to {end_dt.strftime('%Y-%m-%d')}. Work Summary: {pending_work_summary}",
                #         is_read=False
                #     )
                #     session.add(backup_notif)

                session.commit()

                return json.dumps(self.model_to_dict(new_request), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("create_leave_request", e)

    def get_employee_leave_requests(self, employee_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                records = session.query(LeaveRequest).filter_by(employee_id=employee_id).all()
                return json.dumps([self.model_to_dict(r) for r in records], indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_employee_leave_requests", e)

    def get_backup_coverages_by_employee(self, employee_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                records = session.query(LeaveRequest).filter_by(backup_employee_id=employee_id, status='Approved').all()
                res_list = []
                for r in records:
                    r_dict = self.model_to_dict(r)
                    emp = session.query(Employees).filter_by(id=r.employee_id).first()
                    r_dict['employee_name'] = emp.full_name if emp else "Employee"
                    res_list.append(r_dict)
                return json.dumps(res_list, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_backup_coverages_by_employee", e)

    def update_leave_request_status(self, leave_id: str, status: str) -> str:
        with self.SessionLocal() as session:
            try:
                record = session.query(LeaveRequest).filter_by(id=leave_id).first()
                if not record:
                    return json.dumps({"error": "Leave request not found."})
                
                old_status = record.status
                record.status = status
                session.commit()
                session.refresh(record)

                # Audit Log
                self.write_audit_log(
                    user_id="SYSTEM",
                    action="UPDATE_LEAVE_STATUS",
                    target_id=leave_id,
                    details={"old_status": old_status, "new_status": status}
                )

                # 10. Notify Employee
                emp_id = record.employee_id
                emp = session.query(Employees).filter_by(id=emp_id).first()
                emp_name = emp.full_name if emp else "Employee"
                
                new_notif = InAppNotification(
                    user_id=emp_id,
                    title=f"Leave Request {status}",
                    message=f"Your leave request from {record.start_date.strftime('%Y-%m-%d')} to {record.end_date.strftime('%Y-%m-%d')} has been {status.lower()}.",
                    is_read=False
                )
                session.add(new_notif)

                if status == "Approved" and record.backup_employee_id and record.backup_employee_id != "N/A":
                    backup_notif = InAppNotification(
                        user_id=record.backup_employee_id,
                        title="Task Handover Confirmed",
                        message=f"Task Handover Confirmed: The leave request for {emp_name} from {record.start_date.strftime('%Y-%m-%d')} to {record.end_date.strftime('%Y-%m-%d')} has been approved. You are officially confirmed as the backup.",
                        is_read=False
                    )
                    session.add(backup_notif)

                session.commit()

                return json.dumps(self.model_to_dict(record), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("update_leave_request_status", e)

    def cancel_leave_request_by_employee(self, leave_id: str, employee_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                record = session.query(LeaveRequest).filter_by(id=leave_id, employee_id=employee_id).first()
                if not record:
                    return json.dumps({"error": "Leave request not found."})
                
                # Allow cancellation requests as long as the leave has not fully ended yet
                today = datetime.now().date()
                leave_end = record.end_date.date()
                if today > leave_end:
                    return json.dumps({"error": "Cannot cancel or request cancellation of a leave request after the leave has ended."})
                
                if record.status == "Pending":
                    # Direct cancel
                    old_status = record.status
                    record.status = "Cancelled"
                    session.commit()
                    self.write_audit_log(
                        user_id=employee_id,
                        action="CANCEL_LEAVE_REQUEST",
                        target_id=leave_id,
                        details={"old_status": old_status, "status": "Cancelled"}
                    )
                    return json.dumps(self.model_to_dict(record), indent=4, default=str)
                elif record.status == "Approved":
                    # Request cancellation
                    old_status = record.status
                    record.status = "Cancellation Requested"
                    session.commit()
                    self.write_audit_log(
                        user_id=employee_id,
                        action="REQUEST_LEAVE_CANCELLATION",
                        target_id=leave_id,
                        details={"old_status": old_status, "status": "Cancellation Requested"}
                    )
                    # Notify admin
                    admins = session.query(Admins).all()
                    emp = session.query(Employees).filter_by(id=employee_id).first()
                    emp_name = emp.full_name if emp else "Employee"
                    for admin in admins:
                        new_notif = InAppNotification(
                            user_id=admin.id,
                            title="Leave Cancellation Requested",
                            message=f"{emp_name} has requested to cancel their approved leave from {record.start_date.strftime('%Y-%m-%d')} to {record.end_date.strftime('%Y-%m-%d')}.",
                            is_read=False
                        )
                        session.add(new_notif)
                    session.commit()
                    return json.dumps(self.model_to_dict(record), indent=4, default=str)
                else:
                    return json.dumps({"error": f"Cannot cancel a leave request in {record.status} status."})
            except Exception as e:
                session.rollback()
                return self._handle_error("cancel_leave_request_by_employee", e)

    def get_user_notifications(self, user_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                from datetime import datetime, timedelta
                from src.database.database_tables import Employees, ProjectAssignments, Projects, DeveloperTasks, ContentCreatorTasks, ProjectTimeline, InAppNotification, SRS_Documents
                
                thirty_days_ago = datetime.now() - timedelta(days=30)
                seven_days_ago = datetime.now() - timedelta(days=7)
                
                session.query(InAppNotification).filter(
                    InAppNotification.user_id == user_id,
                    InAppNotification.created_at < thirty_days_ago
                ).delete(synchronize_session=False)
                
                session.query(InAppNotification).filter(
                    InAppNotification.user_id == user_id,
                    InAppNotification.is_read == True,
                    InAppNotification.created_at < seven_days_ago
                ).delete(synchronize_session=False)
                
                # --- AUTO-SYNC RISK NOTIFICATIONS FOR EMPLOYEES ---
                # Check if this user is an employee
                employee = session.query(Employees).filter_by(id=user_id, is_active=True).first()
                if employee:
                    # A. Check Projects at Risk
                    # Get assigned projects (via ProjectAssignments)
                    assigned_projects = session.query(Projects).join(
                        ProjectAssignments, Projects.id == ProjectAssignments.project_id
                    ).filter(
                        ProjectAssignments.employee_id == user_id,
                        func.lower(func.coalesce(Projects.status, '')) != 'completed'
                    ).all()
                    
                    # Pre-calculate project accumulated costs
                    dev_costs = {row.project_id: float(row.cost or 0) for row in session.query(
                        DeveloperTasks.project_id, func.sum(DeveloperTasks.employee_cost).label('cost')
                    ).group_by(DeveloperTasks.project_id).all()}
                    
                    content_costs = {row.project_id: float(row.cost or 0) for row in session.query(
                        ContentCreatorTasks.project_id, func.sum(ContentCreatorTasks.employee_cost).label('cost')
                    ).group_by(ContentCreatorTasks.project_id).all()}
                    
                    now = datetime.now()
                    
                    for proj in assigned_projects:
                        is_at_risk = False
                        risk_reasons = []
                        
                        # Check status
                        p_status = (proj.status or "").lower()
                        if "risk" in p_status or "delayed" in p_status or "critical" in p_status:
                            is_at_risk = True
                            risk_reasons.append(f"Status is '{proj.status}'")
                            
                        # Check budget overrun
                        p_budget = float(proj.budget) if proj.budget else 0.0
                        total_cost = dev_costs.get(proj.id, 0.0) + content_costs.get(proj.id, 0.0)
                        if p_budget > 0.0 and total_cost >= p_budget:
                            is_at_risk = True
                            risk_reasons.append(f"Budget overrun (${total_cost:.2f} >= ${p_budget:.2f})")
                        elif p_budget > 0.0 and total_cost >= (p_budget * 0.9):
                            is_at_risk = True
                            risk_reasons.append(f"Nearing budget limit ({total_cost/p_budget*100:.1f}%)")
                            
                        # Check deadline passed
                        if proj.end_date and proj.end_date != "N/A":
                            try:
                                p_end = datetime.strptime(proj.end_date, "%Y-%m-%d")
                                if p_end < now:
                                    is_at_risk = True
                                    risk_reasons.append("Project deadline has passed")
                            except:
                                pass
                                
                        # Check missing SRS
                        srs_exists = session.query(SRS_Documents).filter_by(project_id=proj.id).first()
                        if not srs_exists:
                            is_at_risk = True
                            risk_reasons.append("SRS Document is missing")
                            
                        if is_at_risk:
                            reason_str = ", ".join(risk_reasons)
                            title = f"Project at Risk: {proj.name}"
                            msg = f"Project '{proj.name}' is flagged as at risk due to: {reason_str}."
                            
                            # Check if notification already exists
                            exists = session.query(InAppNotification).filter_by(
                                user_id=user_id,
                                title=title,
                                is_read=False
                            ).first()
                            
                            if not exists:
                                new_notif = InAppNotification(
                                    user_id=user_id,
                                    title=title,
                                    message=msg,
                                    is_read=False
                                )
                                session.add(new_notif)
                                
                    # B. Check Milestones at Risk/Delayed
                    # Get assigned milestones (via MilestoneAssignments)
                    assigned_milestones = session.query(ProjectTimeline, Projects.name).join(
                        MilestoneAssignments, ProjectTimeline.id == MilestoneAssignments.milestone_id
                    ).join(
                        Projects, ProjectTimeline.project_id == Projects.id
                    ).filter(
                        MilestoneAssignments.employee_id == user_id,
                        func.lower(func.coalesce(ProjectTimeline.status, '')) != 'completed'
                    ).all()
                    
                    for milestone, proj_name in assigned_milestones:
                        is_m_at_risk = False
                        m_reasons = []
                        
                        m_status = (milestone.status or "").lower()
                        if m_status == "delayed" or m_status == "at risk":
                            is_m_at_risk = True
                            m_reasons.append(f"Status is '{milestone.status}'")
                            
                        # Check deadline
                        if milestone.expected_end and milestone.expected_end < now:
                            is_m_at_risk = True
                            m_reasons.append("Milestone deadline has passed")
                            
                        if is_m_at_risk:
                            reason_str = ", ".join(m_reasons)
                            title = f"Milestone Delayed: {milestone.milestone_name}"
                            msg = f"Milestone '{milestone.milestone_name}' in project '{proj_name}' is delayed or at risk. Details: {reason_str}."
                            
                            exists = session.query(InAppNotification).filter_by(
                                user_id=user_id,
                                title=title,
                                is_read=False
                            ).first()
                            
                            if not exists:
                                new_notif = InAppNotification(
                                    user_id=user_id,
                                    title=title,
                                    message=msg,
                                    is_read=False
                                )
                                session.add(new_notif)
                                
                    session.commit()
                # ----------------------------------------------------
                
                total_notifs = session.query(InAppNotification).filter_by(user_id=user_id).order_by(InAppNotification.created_at.desc()).all()
                if len(total_notifs) > 50:
                    excess_ids = [n.id for n in total_notifs[50:]]
                    session.query(InAppNotification).filter(InAppNotification.id.in_(excess_ids)).delete(synchronize_session=False)
                
                session.commit()
                
                records = session.query(InAppNotification).filter_by(user_id=user_id).order_by(InAppNotification.created_at.desc()).all()
                return json.dumps([self.model_to_dict(r) for r in records], indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("get_user_notifications", e)

    def mark_notification_as_read(self, notif_id: str, user_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                record = session.query(InAppNotification).filter_by(id=notif_id, user_id=user_id).first()
                if not record:
                    return json.dumps({"error": "Notification not found."})
                record.is_read = True
                session.commit()
                return json.dumps(self.model_to_dict(record), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("mark_notification_as_read", e)

    def get_company_holidays(self) -> str:
        with self.SessionLocal() as session:
            try:
                from src.database.database_tables import CompanyHoliday
                records = session.query(CompanyHoliday).order_by(CompanyHoliday.date.asc()).all()
                return json.dumps([self.model_to_dict(r) for r in records], indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_company_holidays", e)

    def create_company_holiday(self, name: str, date: str, holiday_type: str) -> str:
        with self.SessionLocal() as session:
            try:
                from src.database.database_tables import CompanyHoliday
                new_holiday = CompanyHoliday(
                    name=name,
                    date=date,
                    holiday_type=holiday_type
                )
                session.add(new_holiday)
                session.commit()
                session.refresh(new_holiday)
                return json.dumps(self.model_to_dict(new_holiday), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("create_company_holiday", e)

    def delete_company_holiday(self, holiday_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                from src.database.database_tables import CompanyHoliday
                record = session.query(CompanyHoliday).filter_by(id=holiday_id).first()
                if not record:
                    return json.dumps({"error": "Holiday not found."})
                session.delete(record)
                session.commit()
                return json.dumps({"success": True, "message": "Holiday deleted successfully."})
            except Exception as e:
                session.rollback()
                return self._handle_error("delete_company_holiday", e)

    def bulk_update_working_hours(self, working_hours: float, employee_ids: list = None, department: str = None, role_id: str = None, shift_start_time: str = None, shift_end_time: str = None) -> str:
        with self.SessionLocal() as session:
            try:
                from src.database.database_tables import Employees, Departments_Roles
                query = session.query(Employees)
                if employee_ids:
                    query = query.filter(Employees.id.in_(employee_ids))
                elif role_id:
                    query = query.filter(Employees.role_id == role_id)
                elif department:
                    query = query.join(Departments_Roles, Employees.role_id == Departments_Roles.id).filter(Departments_Roles.department_name == department)
                
                count = 0
                for emp in query.all():
                    emp.working_hours = working_hours
                    if shift_start_time is not None:
                        emp.shift_start_time = shift_start_time
                    if shift_end_time is not None:
                        emp.shift_end_time = shift_end_time
                    count += 1
                session.commit()
                return json.dumps({"success": True, "message": f"Updated working hours to {working_hours} for {count} employees."})
            except Exception as e:
                session.rollback()
                return self._handle_error("bulk_update_working_hours", e)

    def get_employee_assigned_milestones(self, employee_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                from src.database.database_tables import MilestoneAssignments, ProjectTimeline, Projects
                # Pre-calculate project costs
                dev_costs = {row.project_id: float(row.cost or 0) for row in session.query(
                    DeveloperTasks.project_id, func.sum(DeveloperTasks.employee_cost).label('cost')
                ).group_by(DeveloperTasks.project_id).all()}
                
                content_costs = {row.project_id: float(row.cost or 0) for row in session.query(
                    ContentCreatorTasks.project_id, func.sum(ContentCreatorTasks.employee_cost).label('cost')
                ).group_by(ContentCreatorTasks.project_id).all()}

                query = session.query(
                    ProjectTimeline.id.label("milestone_id"),
                    ProjectTimeline.milestone_name.label("milestone_title"),
                    ProjectTimeline.status.label("milestone_status"),
                    ProjectTimeline.expected_start.label("expected_start"),
                    ProjectTimeline.expected_end.label("expected_end"),
                    ProjectTimeline.actual_start.label("actual_start"),
                    ProjectTimeline.actual_end.label("actual_end"),
                    ProjectTimeline.remarks.label("remarks"),
                    ProjectTimeline.sprint_name.label("sprint_name"),
                    ProjectTimeline.module_name.label("module_name"),
                    ProjectTimeline.feature_name.label("feature_name"),
                    ProjectTimeline.work_type.label("work_type"),
                    ProjectTimeline.repo_name.label("repo_name"),
                    Projects.id.label("project_id"),
                    Projects.name.label("project_name"),
                    Projects.budget.label("project_budget"),
                    Projects.end_date.label("project_end_date"),
                    Projects.status.label("project_status")
                ).join(
                    MilestoneAssignments, ProjectTimeline.id == MilestoneAssignments.milestone_id
                ).join(
                    Projects, ProjectTimeline.project_id == Projects.id
                ).filter(
                    MilestoneAssignments.employee_id == employee_id
                ).order_by(ProjectTimeline.expected_start.asc())
                
                now = datetime.now()
                results = []
                for row in query.all():
                    # Calculate project risk
                    proj_cost = dev_costs.get(row.project_id, 0.0) + content_costs.get(row.project_id, 0.0)
                    proj_budget = float(row.project_budget) if row.project_budget else 0.0
                    proj_at_risk = False
                    proj_risk_reasons = []
                    
                    p_status = (row.project_status or "").lower()
                    if "risk" in p_status or "delayed" in p_status or "critical" in p_status:
                        proj_at_risk = True
                        proj_risk_reasons.append(f"Status: {row.project_status}")
                    if proj_budget > 0.0 and proj_cost >= proj_budget:
                        proj_at_risk = True
                        proj_risk_reasons.append("Budget Exceeded")
                    elif proj_budget > 0.0 and proj_cost >= (proj_budget * 0.9):
                        proj_at_risk = True
                        proj_risk_reasons.append("Nearing Budget Limit")
                    if row.project_end_date and row.project_end_date != "N/A":
                        try:
                            p_end = datetime.strptime(row.project_end_date, "%Y-%m-%d")
                            if p_end < now:
                                proj_at_risk = True
                                proj_risk_reasons.append("Deadline Passed")
                        except:
                            pass
                            
                    # Milestone risk
                    is_delayed = False
                    m_status = (row.milestone_status or "").lower()
                    if m_status == "delayed" or m_status == "at risk":
                        is_delayed = True
                    if row.expected_end and row.expected_end < now and m_status != "completed":
                        is_delayed = True
                        
                    results.append({
                        "id": row.milestone_id,
                        "milestone_name": row.milestone_title,
                        "status": row.milestone_status,
                        "expected_start": str(row.expected_start) if row.expected_start else None,
                        "expected_end": str(row.expected_end) if row.expected_end else None,
                        "actual_start": str(row.actual_start) if row.actual_start else None,
                        "actual_end": str(row.actual_end) if row.actual_end else None,
                        "remarks": row.remarks,
                        "sprint_name": row.sprint_name,
                        "module_name": row.module_name,
                        "feature_name": row.feature_name,
                        "work_type": row.work_type,
                        "repo_name": row.repo_name,
                        "project_id": row.project_id,
                        "projectName": row.project_name,
                        "is_delayed": is_delayed,
                        "project_at_risk": proj_at_risk,
                        "project_risk_reasons": proj_risk_reasons
                    })
                return json.dumps(results, indent=4)
            except Exception as e:
                return self._handle_error("get_employee_assigned_milestones", e)

    def edit_project_expense(self, expense_id: str, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                from src.database.database_tables import ProjectExpenses
                expense = session.query(ProjectExpenses).filter_by(id=expense_id).first()
                if not expense:
                    return json.dumps({"error": f"Expense not found."})
                for key, value in data.items():
                    setattr(expense, key, value)
                session.commit()
                session.refresh(expense)
                return json.dumps(self.model_to_dict(expense), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("edit_project_expense", e, context={"expense_id": expense_id, "data": data})

    def get_all_project_assignments(self) -> str:
        with self.SessionLocal() as session:
            try:
                from src.database.database_tables import ProjectAssignments, Employees
                query = session.query(
                    ProjectAssignments.project_id,
                    Employees.full_name
                ).join(
                    Employees, ProjectAssignments.employee_id == Employees.id
                )
                
                mapping = {}
                for row in query.all():
                    if row.project_id not in mapping:
                        mapping[row.project_id] = []
                    mapping[row.project_id].append(row.full_name)
                return json.dumps(mapping, indent=4)
            except Exception as e:
                return self._handle_error("get_all_project_assignments", e)

    # ==========================================
    #           CLIENT RECEIVABLES
    # ==========================================

    def add_client_receivable(self, data: dict) -> str:
        with self.SessionLocal() as session:
            try:
                from src.database.database_tables import ClientReceivables
                new_rec = ClientReceivables(**data)
                session.add(new_rec)
                session.commit()
                session.refresh(new_rec)
                return json.dumps(self.model_to_dict(new_rec), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("add_client_receivable", e, context=data)

    def get_client_receivables(self, project_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                from src.database.database_tables import ClientReceivables
                recs = session.query(ClientReceivables).filter_by(project_id=project_id).order_by(ClientReceivables.due_date.asc()).all()
                return json.dumps([self.model_to_dict(r) for r in recs], indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_client_receivables", e, context={"project_id": project_id})

    def delete_client_receivable(self, receivable_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                from src.database.database_tables import ClientReceivables
                rec = session.query(ClientReceivables).filter_by(id=receivable_id).first()
                if not rec:
                    return json.dumps({"error": "Receivable not found."})
                session.delete(rec)
                session.commit()
                return json.dumps({"success": True})
            except Exception as e:
                session.rollback()
                return self._handle_error("delete_client_receivable", e, context={"receivable_id": receivable_id})

    def mark_client_receivable_done(self, receivable_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                from src.database.database_tables import ClientReceivables, ProjectPayments
                import calendar
                
                rec = session.query(ClientReceivables).filter_by(id=receivable_id).first()
                if not rec:
                    return json.dumps({"error": "Receivable not found."})
                
                # Create a payment ledger entry if amount > 0
                if rec.amount > 0:
                    new_payment = ProjectPayments(
                        project_id=rec.project_id,
                        amount=rec.amount,
                        payment_date=datetime.now(),
                        payment_method="Bank Transfer",
                        reference_number="AUTO-RECEIVABLE",
                        remarks=f"Auto-generated payment from Client Receivable reminder: {rec.item_name}"
                    )
                    session.add(new_payment)
                
                if rec.frequency == "Monthly":
                    # Parse current due date and add 1 month
                    try:
                        curr_date = datetime.strptime(rec.due_date, "%Y-%m-%d")
                    except Exception:
                        curr_date = datetime.now()
                    
                    # Custom standard-library logic to add exactly 1 month
                    month = curr_date.month - 1 + 1
                    year = curr_date.year + month // 12
                    month = month % 12 + 1
                    day = min(curr_date.day, calendar.monthrange(year, month)[1])
                    next_date = datetime(year, month, day)
                    
                    rec.due_date = next_date.strftime("%Y-%m-%d")
                    rec.is_done = False
                else:
                    rec.is_done = True
                
                session.commit()
                session.refresh(rec)
                return json.dumps(self.model_to_dict(rec), indent=4, default=str)
            except Exception as e:
                session.rollback()
                return self._handle_error("mark_client_receivable_done", e, context={"receivable_id": receivable_id})

    def get_pending_handover_tasks(self, assignee_id: str, colleague_id: str) -> str:
        with self.SessionLocal() as session:
            try:
                # Query LeaveRequest for approved, pending handovers
                leave_handovers = session.query(LeaveRequest).filter(
                    LeaveRequest.backup_employee_id == assignee_id,
                    LeaveRequest.employee_id == colleague_id,
                    LeaveRequest.status == 'Approved',
                    LeaveRequest.handover_status != 'Completed'
                ).all()

                combined = []
                for lh in leave_handovers:
                    sprint, module, feature = "N/A", "N/A", "N/A"
                    if lh.milestone_id:
                        milestone = session.query(ProjectTimeline).filter_by(id=lh.milestone_id).first()
                        if milestone:
                            sprint = milestone.sprint_name or "N/A"
                            module = milestone.module_name or "N/A"
                            feature = milestone.feature_name or "N/A"
                            
                    combined.append({
                        "id": lh.id,
                        "project_id": lh.project_id or "",
                        "milestone_id": lh.milestone_id or "",
                        "task_type": lh.task_type or "developer",
                        "pending_work_summary": lh.pending_work_summary or "N/A",
                        "task_performed": lh.pending_work_summary or "N/A",
                        "sprint": sprint,
                        "module": module,
                        "feature": feature,
                        "work_type": "Backend" if lh.task_type == "developer" else "Marketing"
                    })

                # Fallback: Query DeveloperTasks and ContentCreatorTasks for compatibility
                dev_tasks = session.query(DeveloperTasks).filter(
                    DeveloperTasks.is_handover == True,
                    DeveloperTasks.handover_for_employee_id == assignee_id,
                    DeveloperTasks.employee_id == colleague_id,
                    DeveloperTasks.task_status != "Completed"
                ).all()
                for t in dev_tasks:
                    d = self.model_to_dict(t)
                    d['task_type'] = 'developer'
                    combined.append(d)

                content_tasks = session.query(ContentCreatorTasks).filter(
                    ContentCreatorTasks.is_handover == True,
                    ContentCreatorTasks.handover_for_employee_id == assignee_id,
                    ContentCreatorTasks.employee_id == colleague_id,
                    ContentCreatorTasks.task_status != "Completed"
                ).all()
                for t in content_tasks:
                    c = self.model_to_dict(t)
                    c['task_type'] = 'content_creator'
                    combined.append(c)

                return json.dumps(combined, indent=4, default=str)
            except Exception as e:
                return self._handle_error("get_pending_handover_tasks", e)











































